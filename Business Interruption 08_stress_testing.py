"""
BI Step 6 – Stress Testing & Portfolio Resilience  |  Valuation: 2175
======================================================================
All parameters derived live from:
  - BI_data_CLEANED.xlsx         (frequency and severity parameters)
  - srcsc-2026-interest-and-inflation.xlsx  (rate term structure)
  - bi_prospective_exposure.xlsx (total exposure; falls back if unavailable)

No values are hardcoded. Everything flows from the same source data
as Steps 3–5, ensuring the stress test is consistent with the baseline.

Framework
---------
  Frequency : Negative Binomial, system-specific (od from historical claim counts)
  Severity  : Lognormal, system-specific MLE from historical claim amounts
  Product   : Design B Standard — ded=1M, occ_limit=15M (Step 5 lead product)
  Rates     : Step 3 term structure derived from rates file; overridden per stress
  Capital   : VaR(99.5%) − E[L], annual basis (Solvency II proxy)
  Pricing   : Target LR=72%, expense=20%, profit=8%

Stress sections
---------------
  1 – Base Case        (reference, no shocks)
  2 – Single Factor    (one assumption shocked at a time)
  3 – Combined         (two or more factors simultaneously)
  4 – Extreme / Tail   (severe shocks, correlated multi-system events)
"""

import numpy  as np
import pandas as pd
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# 0.  FILE PATHS
# ══════════════════════════════════════════════════════════════════════════════
HISTORICAL_PATH = "/Users/alansteny/Downloads/BI_data_CLEANED.xlsx"
RATES_PATH      = "/Users/alansteny/Downloads/srcsc-2026-interest-and-inflation.xlsx"
EXPOSURE_PATH   = "/Users/alansteny/Downloads/bi_prospective_exposure.xlsx"   # optional
OUTPUT_PATH     = "/Users/alansteny/Downloads/bi_step6_stress_testing.xlsx"

VALUATION_YEAR  = 2175
N               = 10        # projection horizon 2175–2185
N_SIM           = 10_000
RNG_SEED        = 42

# Product Design B (Step 5 lead product)
DED      = 1_000_000
OCC_LIM  = 15_000_000
TARGET_LR     = 0.72
EXPENSE_LOAD  = 0.20
CAP_PCTL      = 99.5         # Solvency II VaR proxy

RETURN_PERIODS = {"1-in-10": 90.0, "1-in-20": 95.0, "1-in-50": 98.0, "1-in-100": 99.0}

# Prospective system → historical analogue mapping (data limitation documented)
SYS_MAP = {
    "Helionis Cluster": "Helionis Cluster",   # direct match
    "Bayesia System":   "Epsilon",             # proxy
    "Oryn Delta":       "Zeta",                # proxy
}


# ══════════════════════════════════════════════════════════════════════════════
# 1.  DERIVE ALL PARAMETERS FROM DATA
# ══════════════════════════════════════════════════════════════════════════════
print("Loading historical data...")

freq_raw = pd.read_excel(HISTORICAL_PATH, sheet_name="Frequency_Cleaned")
sev_raw  = pd.read_excel(HISTORICAL_PATH, sheet_name="Severity_Cleaned")

freq_all = (freq_raw[freq_raw["data_quality_flag"] == 0]
            .dropna(subset=["exposure", "claim_count"])
            .copy())
sev_all  = (sev_raw[(sev_raw["data_quality_flag"] == 0) & (sev_raw["claim_amount"] > 0)]
            .dropna(subset=["claim_amount"])
            .copy())

for c in ["claim_count", "exposure"]:
    freq_all[c] = pd.to_numeric(freq_all[c], errors="coerce")
sev_all["claim_amount"] = pd.to_numeric(sev_all["claim_amount"], errors="coerce")
freq_all = freq_all[freq_all["exposure"] > 0].copy()

print(f"  Frequency rows (flag=0): {len(freq_all):,}")
print(f"  Severity  rows (flag=0): {len(sev_all):,}")

# ── System-specific parameters ─────────────────────────────────────────────
SYSTEMS = {}
for prospective_sys, hist_sys in SYS_MAP.items():
    f = freq_all[freq_all["solar_system"] == hist_sys]
    s = sev_all[ sev_all["solar_system"]  == hist_sys]["claim_amount"].values

    fr    = f["claim_count"].sum() / f["exposure"].sum()
    cc    = f["claim_count"].values
    od    = cc.var(ddof=1) / cc.mean()           # overdispersion ratio var/mean
    ln_mu  = np.log(s).mean()                    # MLE log-mean
    ln_sig = np.log(s).std(ddof=1)               # MLE log-std
    ks_p   = stats.kstest(s, "lognorm", args=(ln_sig, 0, np.exp(ln_mu))).pvalue

    SYSTEMS[prospective_sys] = {
        "hist_sys": hist_sys,
        "fr": fr, "od": od,
        "ln_mu": ln_mu, "ln_sig": ln_sig,
        "n_sev": len(s),
        "sev_mean": s.mean(), "sev_std": s.std(ddof=1),
        "ks_p": ks_p,
    }
    print(f"  {prospective_sys}: fr={fr:.6f}  od={od:.4f}  "
          f"ln_mu={ln_mu:.4f}  ln_sig={ln_sig:.4f}  n={len(s):,}  KS_p={ks_p:.3f}")

# ── Prospective exposure (total active units across portfolio) ──────────────
try:
    exp_df = pd.read_excel(EXPOSURE_PATH, sheet_name="Exposure_Dataset")
    exp_df["active_units"] = pd.to_numeric(exp_df["active_units"], errors="coerce")
    EXPOSURE_0 = float(exp_df["active_units"].sum())
    print(f"  Prospective exposure loaded: {EXPOSURE_0:,.2f} active units")
except Exception:
    # Fall back: use per-unit basis (results are per unit; scale manually)
    EXPOSURE_0 = 1.0
    print("  Prospective exposure file unavailable — running per-unit basis")

# ── Rate term structure from rates file ────────────────────────────────────
print("\nLoading rates...")
r = pd.read_excel(RATES_PATH, sheet_name="Sheet1", header=1, skiprows=[2]).iloc[1:]
r.columns = ["year", "inflation", "overnight", "spot_1y", "spot_10y"]
r = r.dropna(subset=["year"]).copy()
r["year"] = r["year"].astype(int)
for c in ["inflation", "overnight", "spot_1y", "spot_10y"]:
    r[c] = pd.to_numeric(r[c])

avg3 = r.tail(3).mean(numeric_only=True)
avg5 = r.tail(5).mean(numeric_only=True)
avgA = r.mean(numeric_only=True)
last = r.iloc[-1]

# Base rates by projection offset t — Step 3 term structure
def _base_rates(t):
    if   t == 0: return last["inflation"], last["spot_1y"]
    elif t <= 2: return avg3["inflation"], avg3["spot_1y"]
    elif t <= 5: return avg5["inflation"], avg5["spot_1y"]
    else:        return avgA["inflation"], avgA["spot_10y"]

BASE_RATES = {t: _base_rates(t) for t in range(N + 1)}

print(f"  Rates file: {r['year'].min()}–{r['year'].max()}  ({len(r)} years)")
print(f"  t0   inf={BASE_RATES[0][0]:.4%}  disc={BASE_RATES[0][1]:.4%}  [1y spot, last observed]")
print(f"  t1-2 inf={BASE_RATES[1][0]:.4%}  disc={BASE_RATES[1][1]:.4%}  [1y spot, last-3yr avg]")
print(f"  t3-5 inf={BASE_RATES[3][0]:.4%}  disc={BASE_RATES[3][1]:.4%}  [1y spot, last-5yr avg]")
print(f"  t6-10 inf={BASE_RATES[6][0]:.4%} disc={BASE_RATES[6][1]:.4%}  [10y spot, full avg — instrument switch]")


# ══════════════════════════════════════════════════════════════════════════════
# 2.  HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def cum_factors(rates_dict):
    """Cumulative inflation and discount factors from a rates dict."""
    ci = cd = 1.0
    inf_f = {0: 1.0}; disc_f = {0: 1.0}
    for t in range(1, N + 1):
        inf_t, disc_t = rates_dict[t]
        ci  *= (1 + inf_t)
        cd  *= 1 / (1 + disc_t)
        inf_f[t]  = ci
        disc_f[t] = cd
    return inf_f, disc_f


def simulate_system(fr, od, ln_mu, ln_sig, inf_f, disc_f,
                    exposure_0, growth=0.0, corr_boost_t=None, rng=None):
    """
    NB frequency + lognormal severity simulation for one system.
    Per-occurrence deductible and limit applied at claim level.
    Returns nom_mat and disc_mat (N_SIM × N+1).
    """
    if rng is None:
        rng = np.random.default_rng(RNG_SEED)
    p_nb = 1.0 / od
    nom_mat  = np.zeros((N_SIM, N + 1))
    disc_mat = np.zeros((N_SIM, N + 1))
    for t in range(N + 1):
        exp_t = exposure_0 * (1 + growth) ** t
        lam_t = fr * exp_t
        if corr_boost_t and t in corr_boost_t:
            lam_t *= corr_boost_t[t]    # correlated shock multiplier in specific years
        r_t    = max(lam_t / (od - 1), 1e-6)
        counts = rng.negative_binomial(n=r_t, p=p_nb, size=N_SIM)
        max_c  = int(counts.max())
        if max_c == 0:
            net = np.zeros(N_SIM)
        else:
            sm   = rng.lognormal(mean=ln_mu, sigma=ln_sig, size=(N_SIM, max_c))
            mask = np.arange(max_c)[None, :] < counts[:, None]
            net_claims = np.clip(
                np.minimum(np.clip(sm - DED, 0, None), OCC_LIM), 0, None)
            net = (net_claims * mask).sum(axis=1)
        nom_mat[:, t]  = net * inf_f[t]
        disc_mat[:, t] = net * inf_f[t] * disc_f[t]
    return nom_mat, disc_mat


def run_scenario(cfg):
    """Run all systems under a given stress config. Returns portfolio-aggregated arrays."""
    rates = cfg["rates"]
    inf_f, disc_f = cum_factors(rates)
    port_nom  = np.zeros((N_SIM, N + 1))
    port_disc = np.zeros((N_SIM, N + 1))
    for sys_name, sp in SYSTEMS.items():
        fr     = sp["fr"]     * cfg["fr_mult"]
        ln_mu  = sp["ln_mu"]  + cfg["ln_mu_shift"]
        ln_sig = sp["ln_sig"] + cfg["ln_sig_shift"]
        od     = sp["od"]     # structural parameter — not shocked
        rng = np.random.default_rng(RNG_SEED)
        nom, disc = simulate_system(
            fr, od, ln_mu, ln_sig, inf_f, disc_f,
            EXPOSURE_0, cfg["growth"], cfg.get("corr_boost_t"), rng)
        port_nom  += nom
        port_disc += disc
    return port_nom, port_disc


def lev_portfolio(fr_mult=1.0, ln_mu_shift=0.0, ln_sig_shift=0.0, seed=0):
    """Portfolio LEV = sum over systems of fr_i × E[min(max(X-ded,0), lim)]."""
    rng2 = np.random.default_rng(seed)
    total = 0.0
    for sp in SYSTEMS.values():
        ln_mu  = sp["ln_mu"]  + ln_mu_shift
        ln_sig = sp["ln_sig"] + ln_sig_shift
        x = rng2.lognormal(mean=ln_mu, sigma=ln_sig, size=200_000)
        lev = np.clip(np.minimum(np.clip(x - DED, 0, None), OCC_LIM), 0, None).mean()
        total += sp["fr"] * fr_mult * lev * EXPOSURE_0
    return total


def summary_stats(arr):
    d = {"mean":   arr.mean(), "median": np.median(arr), "std": arr.std(ddof=1),
         "P75": np.percentile(arr, 75), "P90": np.percentile(arr, 90),
         "P95": np.percentile(arr, 95), "P98": np.percentile(arr, 98),
         "P99": np.percentile(arr, 99), "P99.5": np.percentile(arr, 99.5)}
    for lbl, pct in RETURN_PERIODS.items():
        d[lbl] = np.percentile(arr, pct)
    return d


# ══════════════════════════════════════════════════════════════════════════════
# 3.  SCENARIO DEFINITIONS
# ══════════════════════════════════════════════════════════════════════════════

def shifted_rates(inf_shift=0.0, disc_shift=0.0):
    return {t: (BASE_RATES[t][0] + inf_shift, BASE_RATES[t][1] + disc_shift)
            for t in range(N + 1)}

def partial_inf_shift(inf_shift, up_to_t):
    """Inflation shock applies only up to year up_to_t, then reverts."""
    return {t: (BASE_RATES[t][0] + (inf_shift if t <= up_to_t else 0.0), BASE_RATES[t][1])
            for t in range(N + 1)}

SCENARIOS = {

    # ── SECTION 1: BASE ──────────────────────────────────────────────────────
    "Base": {
        "section": "1 – Base Case",
        "description": "No shocks. Parameters derived live from BI_data_CLEANED and rates file.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },

    # ── SECTION 2: SINGLE FACTOR ─────────────────────────────────────────────
    "Freq +20%": {
        "section": "2 – Single Factor",
        "description": "Frequency rate × 1.20 across all systems. Severity and rates unchanged.",
        "fr_mult": 1.20, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Freq -20%": {
        "section": "2 – Single Factor",
        "description": "Frequency rate × 0.80. Tests downside / favourable scenario.",
        "fr_mult": 0.80, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Sev mean +25%": {
        "section": "2 – Single Factor",
        "description": "Severity mean +25% via ln_mu shift of log(1.25). Frequency unchanged.",
        "fr_mult": 1.0, "ln_mu_shift": np.log(1.25), "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Sev mean -25%": {
        "section": "2 – Single Factor",
        "description": "Severity mean -25% via ln_mu shift of log(0.75).",
        "fr_mult": 1.0, "ln_mu_shift": np.log(0.75), "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Sev heavier tail": {
        "section": "2 – Single Factor",
        "description": "Severity tail fattened: ln_sig +0.20. More extreme large claims without "
                       "shifting the mean. Represents emergence of a new loss driver.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.20,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Inflation +200bp": {
        "section": "2 – Single Factor",
        "description": "Inflation +200bp across all years. Discount rates unchanged.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": shifted_rates(inf_shift=0.02), "growth": 0.0, "corr_boost_t": None,
    },
    "Inflation -100bp": {
        "section": "2 – Single Factor",
        "description": "Inflation -100bp. Tests benign macro environment.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": shifted_rates(inf_shift=-0.01), "growth": 0.0, "corr_boost_t": None,
    },
    "Disc +200bp": {
        "section": "2 – Single Factor",
        "description": "Discount rates +200bp. Reduces PV of future losses. Inflation unchanged.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": shifted_rates(disc_shift=0.02), "growth": 0.0, "corr_boost_t": None,
    },
    "Disc -200bp": {
        "section": "2 – Single Factor",
        "description": "Discount rates -200bp. Increases PV of future losses.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": shifted_rates(disc_shift=-0.02), "growth": 0.0, "corr_boost_t": None,
    },
    "Exposure +5% pa": {
        "section": "2 – Single Factor",
        "description": "Exposure grows at 5% per annum. Loss parameters unchanged.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.05, "corr_boost_t": None,
    },
    "Exposure -3% pa": {
        "section": "2 – Single Factor",
        "description": "Exposure contracts at 3% per annum.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": -0.03, "corr_boost_t": None,
    },

    # ── SECTION 3: COMBINED ───────────────────────────────────────────────────
    "Freq+Sev up": {
        "section": "3 – Combined",
        "description": "Frequency +20% AND severity mean +25%. Both loss drivers stressed.",
        "fr_mult": 1.20, "ln_mu_shift": np.log(1.25), "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Inf+Freq up": {
        "section": "3 – Combined",
        "description": "Frequency +20% AND inflation +200bp. Cost spiral scenario.",
        "fr_mult": 1.20, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": shifted_rates(inf_shift=0.02), "growth": 0.0, "corr_boost_t": None,
    },
    "Stagflation": {
        "section": "3 – Combined",
        "description": "Inflation +300bp, disc -100bp, frequency +10%. "
                       "Analogue of stagflationary interplanetary macro environment.",
        "fr_mult": 1.10, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": shifted_rates(inf_shift=0.03, disc_shift=-0.01),
        "growth": 0.0, "corr_boost_t": None,
    },
    "Growth+Freq up": {
        "section": "3 – Combined",
        "description": "Exposure +5% pa AND frequency +15%. Rapid expansion with elevated risk.",
        "fr_mult": 1.15, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.05, "corr_boost_t": None,
    },
    "Stress Combined": {
        "section": "3 – Combined",
        "description": "Freq +20%, sev mean +25%, inflation +200bp, disc -100bp, "
                       "exposure +5% pa. All major drivers stressed simultaneously.",
        "fr_mult": 1.20, "ln_mu_shift": np.log(1.25), "ln_sig_shift": 0.0,
        "rates": shifted_rates(inf_shift=0.02, disc_shift=-0.01),
        "growth": 0.05, "corr_boost_t": None,
    },

    # ── SECTION 4: EXTREME / TAIL ─────────────────────────────────────────────
    "Severe inflation spike": {
        "section": "4 – Extreme",
        "description": "Inflation +500bp years 1–5 only, then reverts. "
                       "Sustained interplanetary supply chain disruption.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": partial_inf_shift(0.05, up_to_t=5),
        "growth": 0.0, "corr_boost_t": None,
    },
    "Severe freq spike": {
        "section": "4 – Extreme",
        "description": "Frequency +50%. New technology failure mode or adverse "
                       "environmental conditions across the fleet.",
        "fr_mult": 1.50, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Severe sev spike": {
        "section": "4 – Extreme",
        "description": "Severity mean +50% AND heavier tail (ln_sig +0.30). "
                       "Quantum bore cascade or novel equipment failure class.",
        "fr_mult": 1.0, "ln_mu_shift": np.log(1.50), "ln_sig_shift": 0.30,
        "rates": BASE_RATES, "growth": 0.0, "corr_boost_t": None,
    },
    "Exposure surge": {
        "section": "4 – Extreme",
        "description": "Exposure +15% pa (doubles by yr 5), freq +20%. "
                       "Aggressive expansion outpacing risk management.",
        "fr_mult": 1.20, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.15, "corr_boost_t": None,
    },
    "Correlated solar storm": {
        "section": "4 – Extreme",
        "description": "Solar storm: frequency 3× in years t=3 and t=4 (2178–2179) "
                       "across ALL three systems simultaneously. Correlated BI event.",
        "fr_mult": 1.0, "ln_mu_shift": 0.0, "ln_sig_shift": 0.0,
        "rates": BASE_RATES, "growth": 0.0,
        "corr_boost_t": {3: 3.0, 4: 3.0},
    },
    "Catastrophic: all factors": {
        "section": "4 – Extreme",
        "description": "1-in-100 analogue: freq +50%, sev +50% mean & tail +0.30, "
                       "inflation +400bp, correlated surge year t=5 (freq ×4), exposure +8% pa.",
        "fr_mult": 1.50, "ln_mu_shift": np.log(1.50), "ln_sig_shift": 0.30,
        "rates": shifted_rates(inf_shift=0.04),
        "growth": 0.08, "corr_boost_t": {5: 4.0},
    },
}


# ══════════════════════════════════════════════════════════════════════════════
# 4.  RUN ALL SCENARIOS
# ══════════════════════════════════════════════════════════════════════════════
print(f"\nRunning {len(SCENARIOS)} scenarios × {len(SYSTEMS)} systems × {N_SIM:,} iterations...")

sim_results = {}   # {scen_name: (nom_mat, disc_mat)}  — portfolio-level
for scen_name, cfg in SCENARIOS.items():
    nom, disc = run_scenario(cfg)
    sim_results[scen_name] = (nom, disc)
    print(f"  ✓ {scen_name}")


# ══════════════════════════════════════════════════════════════════════════════
# 5.  AGGREGATE METRICS
# ══════════════════════════════════════════════════════════════════════════════
base_nom_s  = summary_stats(sim_results["Base"][0].sum(axis=1))
base_disc_s = summary_stats(sim_results["Base"][1].sum(axis=1))

summary_rows = []
rp_rows      = []

for scen_name, (nom_mat, disc_mat) in sim_results.items():
    tot_nom  = nom_mat.sum(axis=1)
    tot_disc = disc_mat.sum(axis=1)
    ns = summary_stats(tot_nom)
    ds = summary_stats(tot_disc)

    # LEV-based premium under this stress
    cfg = SCENARIOS[scen_name]
    pure_pp  = lev_portfolio(cfg["fr_mult"], cfg["ln_mu_shift"], cfg["ln_sig_shift"])
    y0       = nom_mat[:, 0]
    el_y0    = y0.mean()
    req_cap  = max(np.percentile(y0, CAP_PCTL) - el_y0, 0)
    prem     = pure_pp / TARGET_LR
    net_rev  = prem * (1 - EXPENSE_LOAD) - el_y0
    roc      = net_rev / req_cap if req_cap > 0 else np.nan

    summary_rows.append({
        "scenario":          scen_name,
        "section":           SCENARIOS[scen_name]["section"],
        "nom_mean":          ns["mean"],
        "nom_std":           ns["std"],
        "nom_P90":           ns["P90"],
        "nom_P95":           ns["P95"],
        "nom_P99":           ns["P99"],
        "disc_mean":         ds["mean"],
        "disc_P90":          ds["P90"],
        "disc_P99":          ds["P99"],
        "required_capital":  req_cap,
        "premium":           prem,
        "net_revenue":       net_rev,
        "roc":               roc,
        "nom_mean_Δ%":       (ns["mean"]  - base_nom_s["mean"])  / base_nom_s["mean"],
        "nom_P90_Δ%":        (ns["P90"]   - base_nom_s["P90"])   / base_nom_s["P90"],
        "nom_P99_Δ%":        (ns["P99"]   - base_nom_s["P99"])   / base_nom_s["P99"],
        "disc_mean_Δ%":      (ds["mean"]  - base_disc_s["mean"]) / base_disc_s["mean"],
    })

    for rp_lbl, rp_pct in RETURN_PERIODS.items():
        rp_rows.append({
            "scenario":       scen_name,
            "section":        SCENARIOS[scen_name]["section"],
            "return_period":  rp_lbl,
            "exceedance_pct": rp_pct,
            "nom_loss":       ns[rp_lbl],
            "disc_loss":      ds[rp_lbl],
            "vs_base_Δ%":    (ns[rp_lbl] - base_nom_s[rp_lbl]) / base_nom_s[rp_lbl],
        })

stress_df = pd.DataFrame(summary_rows)
rp_df     = pd.DataFrame(rp_rows)


# ══════════════════════════════════════════════════════════════════════════════
# 6.  ANNUAL STRESS PATHS (selected scenarios)
# ══════════════════════════════════════════════════════════════════════════════
PATH_SCENS = ["Base", "Stress Combined", "Severe freq spike",
              "Severe sev spike", "Correlated solar storm", "Catastrophic: all factors"]

annual_rows = []
for scen_name in PATH_SCENS:
    if scen_name not in sim_results: continue
    nom_mat, disc_mat = sim_results[scen_name]
    inf_f, _ = cum_factors(SCENARIOS[scen_name]["rates"])
    for t in range(N + 1):
        nom_t  = nom_mat[:, t]
        disc_t = disc_mat[:, t]
        annual_rows.append({
            "scenario":        scen_name,
            "year":            VALUATION_YEAR + t,
            "t":               t,
            "cum_inf_factor":  inf_f[t],
            "mean_nom":        nom_t.mean(),
            "median_nom":      np.median(nom_t),
            "P90_nom":         np.percentile(nom_t, 90),
            "P99_nom":         np.percentile(nom_t, 99),
            "mean_disc":       disc_t.mean(),
        })
annual_df = pd.DataFrame(annual_rows)


# ══════════════════════════════════════════════════════════════════════════════
# 7.  SCENARIO DEFINITIONS SHEET
# ══════════════════════════════════════════════════════════════════════════════
def_rows = []
for sn, cfg in SCENARIOS.items():
    r0 = cfg["rates"][0]; r5 = cfg["rates"][5]; r10 = cfg["rates"][10]
    def_rows.append({
        "scenario":         sn,
        "section":          cfg["section"],
        "description":      cfg["description"],
        "fr_multiplier":    cfg["fr_mult"],
        "ln_mu_shift":      cfg["ln_mu_shift"],
        "ln_sig_shift":     cfg["ln_sig_shift"],
        "exposure_growth":  cfg["growth"],
        "inf_t0":           r0[0],
        "disc_t0":          r0[1],
        "inf_t5":           r5[0],
        "disc_t5":          r5[1],
        "inf_t10":          r10[0],
        "disc_t10":         r10[1],
        "corr_boost":       str(cfg.get("corr_boost_t") or "None"),
    })
def_df = pd.DataFrame(def_rows)


# ══════════════════════════════════════════════════════════════════════════════
# 8.  MANAGEMENT TAKEAWAYS
# ══════════════════════════════════════════════════════════════════════════════
top3 = (stress_df[stress_df["scenario"] != "Base"]
        .sort_values("nom_mean_Δ%", ascending=False).head(3))
roc_neg = stress_df[stress_df["roc"] < 0]["scenario"].tolist()
roc_low = stress_df[(stress_df["roc"] >= 0) & (stress_df["roc"] < 0.05)]["scenario"].tolist()

# System-level vulnerability: base P99 vs catastrophic P99
sys_vuln = {}
for sys_name, sp in SYSTEMS.items():
    rng_b = np.random.default_rng(RNG_SEED)
    inf_b, disc_b = cum_factors(BASE_RATES)
    nm_b, _ = simulate_system(sp["fr"], sp["od"], sp["ln_mu"], sp["ln_sig"],
                               inf_b, disc_b, EXPOSURE_0, rng=rng_b)
    cfg_c = SCENARIOS["Catastrophic: all factors"]
    inf_c, disc_c = cum_factors(cfg_c["rates"])
    rng_c = np.random.default_rng(RNG_SEED)
    nm_c, _ = simulate_system(
        sp["fr"] * cfg_c["fr_mult"],
        sp["od"],
        sp["ln_mu"] + cfg_c["ln_mu_shift"],
        sp["ln_sig"] + cfg_c["ln_sig_shift"],
        inf_c, disc_c, EXPOSURE_0, cfg_c["growth"],
        cfg_c.get("corr_boost_t"), rng_c)
    bp99 = np.percentile(nm_b.sum(axis=1), 99)
    cp99 = np.percentile(nm_c.sum(axis=1), 99)
    sys_vuln[sys_name] = {"base_P99": bp99, "cat_P99": cp99,
                           "mult": cp99 / bp99 if bp99 > 0 else np.nan}

most_vuln = max(sys_vuln, key=lambda s: sys_vuln[s]["mult"])

cat_port_p99 = np.percentile(sim_results["Catastrophic: all factors"][0].sum(axis=1), 99)
corr_port_p99 = np.percentile(sim_results["Correlated solar storm"][0].sum(axis=1), 99)

take_rows = [
    ("PARAMETER SOURCE", "", ""),
    ("Historical data", "BI_data_CLEANED.xlsx — flag=0, both sheets", ""),
    ("Rates file", "srcsc-2026-interest-and-inflation.xlsx — derived term structure", ""),
    ("Frequency rows used", f"{len(freq_all):,}", "Negative Binomial per system"),
    ("Severity rows used",  f"{len(sev_all):,}",  "Lognormal per system, KS tested"),
    ("Exposure basis", f"{EXPOSURE_0:,.2f} active units" if EXPOSURE_0 > 1 else "Per-unit (scale by portfolio)", ""),
    ("", "", ""),
    ("STRESS IMPACT RANKING", "", ""),
    *[(f"  #{i+1}: {row['scenario']}", f"Mean +{row['nom_mean_Δ%']:.1%}",
       f"P99 +{row['nom_P99_Δ%']:.1%}")
      for i, (_, row) in enumerate(top3.iterrows())],
    ("", "", ""),
    ("RETURN PERIODS — BASE CASE", "", ""),
    ("1-in-10  (P90)", f"{base_nom_s['1-in-10']:,.0f}", "Attritional planning threshold"),
    ("1-in-20  (P95)", f"{base_nom_s['1-in-20']:,.0f}", "Reinsurance attachment"),
    ("1-in-50  (P98)", f"{base_nom_s['1-in-50']:,.0f}", "Capital buffer calibration"),
    ("1-in-100 (P99)", f"{base_nom_s['1-in-100']:,.0f}", "Extreme capital / stop-loss"),
    ("1-in-100 under Catastrophic", f"{cat_port_p99:,.0f}", "Upper planning bound"),
    ("1-in-100 Correlated storm",   f"{corr_port_p99:,.0f}", "Cross-system event — XL trigger"),
    ("", "", ""),
    ("SYSTEM VULNERABILITY (Catastrophic vs Base P99)", "", ""),
    *[(f"  {s}", f"base={sys_vuln[s]['base_P99']:,.0f}  cat={sys_vuln[s]['cat_P99']:,.0f}",
       f"{sys_vuln[s]['mult']:.1f}× uplift")
      for s in SYSTEMS],
    ("Most vulnerable", most_vuln, f"{sys_vuln[most_vuln]['mult']:.1f}× P99 uplift"),
    ("", "", ""),
    ("CAPITAL & PRICING ADEQUACY", "", ""),
    ("Negative ROC scenarios",  str(roc_neg)  if roc_neg  else "None", "Premium structurally inadequate"),
    ("ROC < 5% scenarios",      str(roc_low[:4]) if roc_low else "None", "Repricing recommended"),
    ("", "", ""),
    ("RECOMMENDED MANAGEMENT ACTIONS", "", ""),
    ("Reinsurance", "XL cover above 1-in-20 per system per annum", "Protects under correlated events"),
    ("Repricing trigger", "If realised freq exceeds base +15% for 2+ years", "Or inflation sustained +200bp"),
    ("Capital review", "If correlated storm or catastrophic scenario materialises", "Immediate capital call"),
    ("Monitoring priority", f"{most_vuln} — heaviest tail risk", "Review annually"),
    ("BI data limitation", "Bayesia←Epsilon proxy; Oryn Delta←Zeta proxy", "Validate as experience develops"),
]
take_df = pd.DataFrame(take_rows, columns=["item", "value", "notes"])


# ══════════════════════════════════════════════════════════════════════════════
# 9.  WRITE WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
DB   = PatternFill("solid", start_color="1F4E79")
MB   = PatternFill("solid", start_color="2E75B6")
GB   = PatternFill("solid", start_color="375623")
HF   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BF   = Font(name="Arial", size=10)
thin = Side(style="thin", color="BFBFBF")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

SEC_FILL = {
    "1 – Base Case":    PatternFill("solid", start_color="D6E4F0"),
    "2 – Single Factor":PatternFill("solid", start_color="E9F7EF"),
    "3 – Combined":     PatternFill("solid", start_color="FEF9E7"),
    "4 – Extreme":      PatternFill("solid", start_color="FDEDEC"),
}

def aw(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max((len(str(c.value or "")) for c in col), default=8) + 3, 52)

def ws_write(ws, df, money=None, pct=None, section_col=None, hdr_fill=None):
    money = money or []; pct = pct or []
    hfill = hdr_fill or DB
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(1, c, col)
        cell.font, cell.fill = HF, hfill
        cell.alignment, cell.border = Alignment(horizontal="center"), BDR
    for ri, row_t in enumerate(df.itertuples(index=False), 2):
        sec  = row_t[df.columns.get_loc(section_col)] if section_col and section_col in df.columns else None
        rfill = SEC_FILL.get(sec)
        item = row_t[df.columns.get_loc("item")] if "item" in df.columns else ""
        is_hdr = str(item).isupper() and str(item).strip() != ""
        for c, val in enumerate(row_t, 1):
            cell = ws.cell(ri, c, val if pd.notna(val) else "")
            cell.border = BDR
            cname = df.columns[c - 1]
            if is_hdr:
                cell.font, cell.fill = HF, MB
            else:
                cell.font = BF
                if rfill: cell.fill = rfill
            if not is_hdr:
                if cname in money: cell.number_format = "#,##0"
                if cname in pct:   cell.number_format = "0.00%"
    aw(ws)

M = ["nom_mean","nom_std","nom_P90","nom_P95","nom_P99",
     "disc_mean","disc_P90","disc_P99","required_capital","premium","net_revenue",
     "nom_loss","disc_loss","mean_nom","median_nom","P90_nom","P99_nom","mean_disc"]
P = ["roc","nom_mean_Δ%","nom_P90_Δ%","nom_P99_Δ%","disc_mean_Δ%","vs_base_Δ%",
     "fr_multiplier","exposure_growth","inf_t0","disc_t0","inf_t5","disc_t5","inf_t10","disc_t10"]

ws1 = wb.active; ws1.title = "stress_summary"
ws_write(ws1, stress_df, money=M, pct=P, section_col="section"); ws1.freeze_panes = "C2"

ws2 = wb.create_sheet("return_periods")
ws_write(ws2, rp_df, money=M, pct=P, section_col="section"); ws2.freeze_panes = "C2"

ws3 = wb.create_sheet("scenario_definitions")
ws_write(ws3, def_df, money=M, pct=P)

ws4 = wb.create_sheet("annual_stress_paths")
ws_write(ws4, annual_df, money=M, pct=P); ws4.freeze_panes = "B2"

ws5 = wb.create_sheet("management_takeaways")
ws_write(ws5, take_df, hdr_fill=GB, section_col="item")

wb.save(OUTPUT_PATH)


# ══════════════════════════════════════════════════════════════════════════════
# 10. CONSOLE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
print()
print("=" * 72)
print("  BI STEP 6 – STRESS TESTING SUMMARY")
print(f"  Data: BI_data_CLEANED + srcsc rates  |  N_sim={N_SIM:,}  |  {len(SCENARIOS)} scenarios")
print(f"  Product Design B: ded=1M  lim=15M  Target LR={TARGET_LR:.0%}")
print("=" * 72)
print(f"\n  {'Scenario':<36} {'Mean':>14} {'ΔMean':>7} {'P99':>14} {'ΔP99':>7}")
for _, row in stress_df.iterrows():
    tag = " ◄" if row["section"] == "4 – Extreme" else ""
    print(f"  {row['scenario']:<36} {row['nom_mean']:>14,.0f} "
          f"{row['nom_mean_Δ%']:>7.1%} {row['nom_P99']:>14,.0f} "
          f"{row['nom_P99_Δ%']:>7.1%}{tag}")
print()
print("  RETURN PERIODS (base, portfolio nominal 10yr):")
for lbl, pct in RETURN_PERIODS.items():
    print(f"    {lbl} (P{pct:.0f}): {base_nom_s[lbl]:>14,.0f}")
print()
print("  KEY RISK DRIVERS:")
for i, (_, row) in enumerate(top3.iterrows(), 1):
    print(f"    {i}. {row['scenario']}  (mean +{row['nom_mean_Δ%']:.1%}, P99 +{row['nom_P99_Δ%']:.1%})")
print(f"\n  MOST VULNERABLE SYSTEM: {most_vuln}  ({sys_vuln[most_vuln]['mult']:.1f}× P99 under Catastrophic)")
if roc_neg:
    print(f"  CAPITAL ALERT: negative ROC under — {roc_neg}")
else:
    print("  Capital: no negative ROC scenarios under Design B pricing")
print()
print(f"  Output → {OUTPUT_PATH}")
print("=" * 72)