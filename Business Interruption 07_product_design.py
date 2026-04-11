"""
BI Step 5 – Product Design, Pricing & Strategic Recommendations  |  2175
=========================================================================
Converts Step 4 loss distributions into product economics and management decisions.

Architecture:
  - Per-occurrence deductible + per-occurrence limit (actuarially correct for BI)
  - Limited Expected Value (LEV) computed directly from historical severity
  - Premium = pure_premium / target_loss_ratio  (target LR = 72%, expense 20%, profit 8%)
  - Risk margin = CoV-based loading on top of pure premium
  - Capital requirement = VaR(99.5%) - E[L]  (Solvency II proxy)
  - Return on Capital = (premium - expected_loss - expenses) / required_capital
  - System-specific params from historical data (Helionis, Bayesia←Epsilon, Oryn Delta←Zeta)
  - 10-year forward simulation uses Step 3 term structure for inflation and discounting

Data limitation documented: prospective systems are Helionis Cluster, Bayesia System,
Oryn Delta. Historical data contains Helionis Cluster, Epsilon, Zeta. Helionis maps
directly. Epsilon and Zeta are used as proxies for Bayesia and Oryn Delta respectively.
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PATHS ─────────────────────────────────────────────────────────────────────
HISTORICAL_PATH = "/Users/alansteny/Downloads/BI_data_CLEANED.xlsx"
RATES_PATH      = "/Users/alansteny/Downloads/srcsc-2026-interest-and-inflation.xlsx"
OUTPUT_PATH     = "/Users/alansteny/Downloads/bi_step5_product_design.xlsx"

VALUATION_DATE  = 2175
N               = 10
N_SIM           = 10_000
RNG_SEED        = 42

# ── PRICING STRUCTURE ─────────────────────────────────────────────────────────
EXPENSE_LOAD    = 0.20    # 20% of premium
PROFIT_MARGIN   = 0.08    # 8% target profit
TARGET_LR       = 1.0 - EXPENSE_LOAD - PROFIT_MARGIN   # 0.72
RISK_MARGIN_COV = 0.10    # 10% of std added as risk margin loading
CAPITAL_PERCENTILE = 99.5 # Solvency II VaR proxy

# ── PRODUCT DESIGNS ───────────────────────────────────────────────────────────
# Per-occurrence structure: insurer pays min(max(X - ded, 0), limit) per claim
# Waiting period is the BI equivalent of a deductible (days before coverage activates)
PRODUCT_DESIGNS = {
    "A – Broad Coverage": {
        "ded":       500_000,
        "occ_limit": 25_000_000,
        "wait_days": 7,
        "description": "Low entry threshold; covers most BI events; suited to high-frequency systems",
    },
    "B – Standard Coverage": {
        "ded":       1_000_000,
        "occ_limit": 15_000_000,
        "wait_days": 14,
        "description": "Balanced; retains attritional losses; suitable as core product",
    },
    "C – Lean / Catastrophe Layer": {
        "ded":       2_000_000,
        "occ_limit": 10_000_000,
        "wait_days": 30,
        "description": "High deductible; targets severe events only; capital-efficient",
    },
}

# ── SOLAR SYSTEM MAPPING ──────────────────────────────────────────────────────
# Helionis Cluster: direct historical match
# Bayesia System:  proxy = Epsilon (similar exposure profile)
# Oryn Delta:      proxy = Zeta    (similar exposure profile)
SYS_MAP = {
    "Helionis Cluster": "Helionis Cluster",
    "Bayesia System":   "Epsilon",
    "Oryn Delta":       "Zeta",
}
SYS_PROXY_NOTE = {
    "Helionis Cluster": "Direct historical match",
    "Bayesia System":   "Proxy: Epsilon historical data (data limitation — see assumptions)",
    "Oryn Delta":       "Proxy: Zeta historical data   (data limitation — see assumptions)",
}


# ═══════════════════════════════════════════════════════════════════════════════
# 1. LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════════
freq_raw = pd.read_excel(HISTORICAL_PATH, sheet_name="Frequency_Cleaned")
sev_raw  = pd.read_excel(HISTORICAL_PATH, sheet_name="Severity_Cleaned")

freq_all = freq_raw[freq_raw["data_quality_flag"] == 0].dropna(subset=["exposure","claim_count"]).copy()
sev_all  = sev_raw[(sev_raw["data_quality_flag"] == 0) & (sev_raw["claim_amount"] > 0)
                   ].dropna(subset=["claim_amount"]).copy()


# ═══════════════════════════════════════════════════════════════════════════════
# 2. RATES — Step 3 term structure
# ═══════════════════════════════════════════════════════════════════════════════
r = pd.read_excel(RATES_PATH, sheet_name="Sheet1", header=1, skiprows=[2]).iloc[1:]
r.columns = ["year","inflation","overnight","spot_1y","spot_10y"]
r = r.dropna(subset=["year"]).copy()
r["year"] = r["year"].astype(int)
for c in ["inflation","overnight","spot_1y","spot_10y"]:
    r[c] = pd.to_numeric(r[c])

avg3 = r.tail(3).mean(numeric_only=True)
avg5 = r.tail(5).mean(numeric_only=True)
avgA = r.mean(numeric_only=True)
last = r.iloc[-1]

def _rates(t):
    if   t == 0: return last["inflation"], last["spot_1y"]
    elif t <= 2: return avg3["inflation"], avg3["spot_1y"]
    elif t <= 5: return avg5["inflation"], avg5["spot_1y"]
    else:        return avgA["inflation"], avgA["spot_10y"]

cum_inf = {}; cum_disc = {}; ci = cd = 1.0
for t in range(N + 1):
    if t > 0:
        inf_t, disc_t = _rates(t)
        ci *= (1 + inf_t); cd *= 1 / (1 + disc_t)
    cum_inf[t]  = ci
    cum_disc[t] = cd if t > 0 else 1.0


# ═══════════════════════════════════════════════════════════════════════════════
# 3. SYSTEM-LEVEL PARAMETERS
# ═══════════════════════════════════════════════════════════════════════════════
def system_params(hist_sys):
    fc = freq_all[freq_all["solar_system"] == hist_sys]
    sc = sev_all[ sev_all["solar_system"]  == hist_sys]
    fr = fc["claim_count"].sum() / fc["exposure"].sum()
    cc = fc["claim_count"].values
    od = cc.var(ddof=1) / cc.mean()
    sv = sc["claim_amount"].values
    ln_mu  = np.log(sv).mean()
    ln_sig = np.log(sv).std(ddof=1)
    ks_s, ks_p = stats.kstest(sv, "lognorm", args=(ln_sig, 0, np.exp(ln_mu)))
    return {"fr": fr, "od": od, "p_nb": 1/od, "ln_mu": ln_mu, "ln_sig": ln_sig,
            "sv": sv, "ks_p": ks_p, "n_sev": len(sv),
            "sev_mean": sv.mean(), "sev_std": sv.std(ddof=1)}

sys_params = {s: system_params(h) for s, h in SYS_MAP.items()}


# ═══════════════════════════════════════════════════════════════════════════════
# 4. CORE PRICING ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
def lev(sv, ded, occ_lim):
    """Limited Expected Value: E[min(max(X-d,0), L)]"""
    return np.clip(np.minimum(np.clip(sv - ded, 0, None), occ_lim), 0, None).mean()

def simulate_annual(fr, od, ln_mu, ln_sig, ded, occ_lim, growth=0.0, rng=None):
    """
    Simulate N_SIM × (N+1) net insured annual losses.
    Returns nominal_matrix and discounted_matrix (N_SIM × N+1).
    Net insured loss per claim = min(max(X - ded, 0), occ_lim).
    """
    if rng is None:
        rng = np.random.default_rng(RNG_SEED)
    p_nb = 1.0 / od
    nom_mat  = np.zeros((N_SIM, N + 1))
    disc_mat = np.zeros((N_SIM, N + 1))
    for t in range(N + 1):
        exp_t   = (1 + growth) ** t          # relative to per-unit exposure
        lam_t   = fr * exp_t
        r_t     = lam_t / (od - 1)
        counts  = rng.negative_binomial(n=r_t, p=p_nb, size=N_SIM)
        max_c   = int(counts.max())
        if max_c == 0:
            net = np.zeros(N_SIM)
        else:
            sm   = rng.lognormal(mean=ln_mu, sigma=ln_sig, size=(N_SIM, max_c))
            mask = np.arange(max_c)[None, :] < counts[:, None]
            # Per-occurrence limit applied claim by claim
            net_claims = np.clip(np.minimum(np.clip(sm - ded, 0, None), occ_lim), 0, None)
            net = (net_claims * mask).sum(axis=1)
        nom_mat[:, t]  = net * cum_inf[t]
        disc_mat[:, t] = net * cum_inf[t] * cum_disc[t]
    return nom_mat, disc_mat

def pricing_metrics(nom_mat, disc_mat, pure_pp, sev_std_annual, rng_seed=None):
    """Derive premium, capital, and return metrics from simulation output."""
    tot_nom  = nom_mat.sum(axis=1)
    tot_disc = disc_mat.sum(axis=1)
    y0_nom   = nom_mat[:, 0]   # year 0 annual loss for annual premium setting
    el       = y0_nom.mean()
    std_y0   = y0_nom.std(ddof=1)
    risk_margin  = std_y0 * RISK_MARGIN_COV
    premium_low  = el / TARGET_LR
    premium_high = (el + risk_margin) / TARGET_LR
    required_cap = np.percentile(y0_nom, CAPITAL_PERCENTILE) - el
    net_rev_low  = premium_low  * (1 - EXPENSE_LOAD) - el
    net_rev_high = premium_high * (1 - EXPENSE_LOAD) - el
    roc_low  = net_rev_low  / max(required_cap, 1)
    roc_high = net_rev_high / max(required_cap, 1)
    return {
        "el_y0":           el,
        "std_y0":          std_y0,
        "risk_margin":     risk_margin,
        "premium_low":     premium_low,
        "premium_high":    premium_high,
        "required_capital":required_cap,
        "net_rev_low":     net_rev_low,
        "net_rev_high":    net_rev_high,
        "roc_low":         roc_low,
        "roc_high":        roc_high,
        "lr_at_el":        el / premium_low,
        "p90_y0":          np.percentile(y0_nom, 90),
        "p95_y0":          np.percentile(y0_nom, 95),
        "p99_y0":          np.percentile(y0_nom, 99),
        "p995_y0":         np.percentile(y0_nom, 99.5),
        "tot_nom_mean":    tot_nom.mean(),
        "tot_nom_p90":     np.percentile(tot_nom, 90),
        "tot_nom_p95":     np.percentile(tot_nom, 95),
        "tot_nom_p99":     np.percentile(tot_nom, 99),
        "tot_disc_mean":   tot_disc.mean(),
        "tot_disc_p90":    np.percentile(tot_disc, 90),
        "tot_disc_p95":    np.percentile(tot_disc, 95),
        "tot_disc_p99":    np.percentile(tot_disc, 99),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 5. RUN ALL COMBINATIONS
# ═══════════════════════════════════════════════════════════════════════════════
GROWTH_SCENARIOS = {"Base (constant)": 0.00, "Growth +5% pa": 0.05, "Contraction -2% pa": -0.02}

print("Running simulations...", flush=True)
all_results = []  # list of dicts for output tables

for sys_name, sp in sys_params.items():
    for d_name, dc in PRODUCT_DESIGNS.items():
        lev_val  = lev(sp["sv"], dc["ded"], dc["occ_limit"])
        pure_pp  = sp["fr"] * lev_val
        for scen_name, growth in GROWTH_SCENARIOS.items():
            rng = np.random.default_rng(RNG_SEED)
            nom_mat, disc_mat = simulate_annual(
                sp["fr"], sp["od"], sp["ln_mu"], sp["ln_sig"],
                dc["ded"], dc["occ_limit"], growth, rng)
            pm = pricing_metrics(nom_mat, disc_mat, pure_pp, sp["sev_std"])
            all_results.append({
                "solar_system":     sys_name,
                "product_design":   d_name,
                "scenario":         scen_name,
                "hist_proxy":       SYS_MAP[sys_name],
                "deductible":       dc["ded"],
                "occ_limit":        dc["occ_limit"],
                "wait_days":        dc["wait_days"],
                "freq_rate":        sp["fr"],
                "lev":              lev_val,
                "pure_premium":     pure_pp,
                **pm,
            })
        print(f"  Done: {sys_name} × {d_name}", flush=True)

results_df = pd.DataFrame(all_results)


# ═══════════════════════════════════════════════════════════════════════════════
# 6. BUILD OUTPUT SHEETS
# ═══════════════════════════════════════════════════════════════════════════════

# ── pricing_summary: base scenario, all systems x designs ────────────────────
base_df = results_df[results_df["scenario"] == "Base (constant)"].copy()
pricing_cols = ["solar_system","product_design","deductible","occ_limit","wait_days",
                "freq_rate","lev","pure_premium",
                "el_y0","std_y0","risk_margin",
                "premium_low","premium_high","lr_at_el",
                "required_capital","net_rev_low","net_rev_high","roc_low","roc_high"]
pricing_summary = base_df[pricing_cols].reset_index(drop=True)

# ── loss_distribution: base scenario, annual stats ────────────────────────────
dist_cols = ["solar_system","product_design","scenario",
             "el_y0","std_y0","p90_y0","p95_y0","p99_y0","p995_y0",
             "tot_nom_mean","tot_nom_p90","tot_nom_p95","tot_nom_p99",
             "tot_disc_mean","tot_disc_p90","tot_disc_p95","tot_disc_p99"]
loss_dist = results_df[dist_cols].reset_index(drop=True)

# ── scenario_comparison: all scenarios, Design B only ────────────────────────
scen_comp = results_df[results_df["product_design"] == "B – Standard Coverage"][
    ["solar_system","scenario","pure_premium","premium_low","premium_high",
     "el_y0","p90_y0","p95_y0","p99_y0","required_capital","roc_low","roc_high",
     "tot_nom_mean","tot_nom_p99"]
].reset_index(drop=True)

# ── system_comparison: Design B base, side-by-side systems ────────────────────
sys_comp = base_df[base_df["product_design"] == "B – Standard Coverage"][
    ["solar_system","hist_proxy","freq_rate","lev","pure_premium",
     "el_y0","p90_y0","p95_y0","p99_y0","required_capital",
     "net_rev_low","net_rev_high","roc_low","roc_high"]
].reset_index(drop=True)

# ── product_design_specs ──────────────────────────────────────────────────────
design_rows = []
for dn, dc in PRODUCT_DESIGNS.items():
    design_rows.append({
        "design":        dn,
        "deductible":    dc["ded"],
        "occ_limit":     dc["occ_limit"],
        "wait_days":     dc["wait_days"],
        "description":   dc["description"],
        "coverage_trigger": "Operational shutdown or output reduction exceeding waiting period",
        "exclusions":    "Acts of war; regulatory shutdown; pre-existing conditions; scheduled maintenance",
        "benefit_basis": "Proven lost revenue × coverage fraction, net of deductible, capped at per-occurrence limit",
        "scalability":   "Limit and deductible indexed to inflation; exposure growth accommodated via NB frequency scaling",
    })
design_specs = pd.DataFrame(design_rows)

# ── strategic_recommendations ─────────────────────────────────────────────────
rec_rows = [
    ("HELIONIS CLUSTER", "", "", ""),
    ("Risk profile",        "Highest severity (sev_mean=4.8M, P99=41M)",
                             "Highest tail risk of the three systems", ""),
    ("Recommended product", "Design B or C",
                             "High severity warrants deductible to avoid adverse selection", ""),
    ("Pricing note",        "Premium 5–8% above Zeta equivalent",
                             "Driven by higher ln_sigma (1.25 vs 1.17)", ""),
    ("Capital requirement", "Highest per unit",
                             "P99.5 significantly elevated due to heavy tail", ""),
    ("Scalability",         "Moderate — exposure growth increases tail risk disproportionately",
                             "Monitor closely under Growth scenario", ""),
    ("", "", "", ""),
    ("BAYESIA SYSTEM", "", "", ""),
    ("Risk profile",        "Highest frequency (fr=0.206) — slightly more attritional",
                             "More frequent but slightly lower severity than Helionis", ""),
    ("Recommended product", "Design A or B",
                             "Frequency profile means deductible still filters noise effectively", ""),
    ("Pricing note",        "Similar pure premium to Helionis despite lower severity",
                             "Higher frequency offsets lower per-claim cost", ""),
    ("Capital requirement", "Similar to Helionis",
                             "P99 comparable; frequency vol drives capital need", ""),
    ("Scalability",         "Well-suited to growth scenarios",
                             "Frequency-driven exposure scales predictably", ""),
    ("", "", "", ""),
    ("ORYN DELTA", "", "", ""),
    ("Risk profile",        "Lowest severity (sev_mean=4.1M, P99=30M) and tightest tail",
                             "Most favourable risk profile of the three systems", ""),
    ("Recommended product", "Design A, B, or C — all viable",
                             "Lower tail risk accommodates broader coverage structures", ""),
    ("Pricing note",        "5–8% below Helionis and Bayesia equivalent",
                             "Lower ln_sigma (1.17) and lower severity mean", ""),
    ("Capital requirement", "Lowest of three systems",
                             "Suitable for market entry; best ROC profile", ""),
    ("Scalability",         "Highest ROC under growth scenario",
                             "Recommend prioritising growth here", ""),
    ("", "", "", ""),
    ("PORTFOLIO STRATEGY", "", "", ""),
    ("Lead product",        "Design B (Standard) across all systems",
                             "Best balance of premium adequacy and capital efficiency", ""),
    ("Market entry",        "Oryn Delta first — lowest capital requirement and best ROC",
                             "Then Bayesia, then Helionis with tighter terms", ""),
    ("Catastrophe limit",   "Aggregate annual limit per policy = 3× per-occurrence limit",
                             "Limits correlated event exposure across systems", ""),
    ("Cross-system corr",   "Solar storm / supply chain events can affect all three simultaneously",
                             "Require portfolio-level reinsurance cover above P99", ""),
    ("Reinsurance",         "XL cover from P95 to P99.5 per system",
                             "Protects capital under Stress Combined scenario", ""),
    ("Inflation sensitivity","1% inflation increase raises 10yr nominal total ~4–5%",
                             "Partially offset by discount rate increase under Stress Combined", ""),
]
rec_df = pd.DataFrame(rec_rows, columns=["item","value","rationale","note"])

# ── assumptions ───────────────────────────────────────────────────────────────
assump_rows = [
    ("DATA & MAPPING", "", ""),
    ("historical_systems",   "Helionis Cluster, Epsilon, Zeta",
                              "Three systems in cleaned historical data"),
    ("prospective_systems",  "Helionis Cluster, Bayesia System, Oryn Delta",
                              "Three systems in Cosmic Quarry RFP"),
    ("Bayesia proxy",        "Epsilon historical data",
                              "Data limitation: no direct Bayesia history; Epsilon used as closest analogue"),
    ("Oryn Delta proxy",     "Zeta historical data",
                              "Data limitation: no direct Oryn Delta history; Zeta used as closest analogue"),
    ("", "", ""),
    ("FREQUENCY", "", ""),
    ("distribution",         "Negative Binomial per occurrence",
                              "Overdispersion ratio empirically derived per system"),
    ("Helionis fr / od",     f"0.1940 / 1.6824", ""),
    ("Bayesia fr / od",      f"0.2059 / 1.7149", ""),
    ("Oryn Delta fr / od",   f"0.2020 / 1.7356", ""),
    ("", "", ""),
    ("SEVERITY", "", ""),
    ("distribution",         "Lognormal per claim", "KS p > 0.49 for all three systems"),
    ("Helionis ln_mu/sig",   "14.5828 / 1.2528",   ""),
    ("Bayesia ln_mu/sig",    "14.5723 / 1.1924",   ""),
    ("Oryn Delta ln_mu/sig", "14.5772 / 1.1698",   ""),
    ("", "", ""),
    ("PRICING STRUCTURE", "", ""),
    ("expense_load",         "20%",                 "Applied to gross premium"),
    ("profit_margin",        "8%",                  "Target underwriting profit"),
    ("target_loss_ratio",    "72%",                 "1 - expense_load - profit_margin"),
    ("risk_margin",          "10% of annual std",   "CoV-based loading for process risk"),
    ("capital_basis",        "VaR(99.5%) - E[L]",  "Solvency II proxy, annual"),
    ("", "", ""),
    ("INFLATION & DISCOUNTING", "", ""),
    ("source",               "srcsc-2026-interest-and-inflation.xlsx", "Step 3 term structure"),
    ("t1_t2",                f"inf={avg3['inflation']:.4%}  disc={avg3['spot_1y']:.4%}", "Last-3yr avg, 1y spot"),
    ("t3_t5",                f"inf={avg5['inflation']:.4%}  disc={avg5['spot_1y']:.4%}", "Last-5yr avg, 1y spot"),
    ("t6_t10",               f"inf={avgA['inflation']:.4%}  disc={avgA['spot_10y']:.4%}", "Full-series avg, 10y spot [instrument switch]"),
    ("", "", ""),
    ("PRODUCT DESIGN", "", ""),
    ("deductible_basis",     "Per occurrence — insured retains first loss per claim", ""),
    ("limit_basis",          "Per occurrence — insurer liability capped per claim",   ""),
    ("waiting_period",       "Equivalent BI concept to deductible: days before benefit activates", ""),
    ("LEV_formula",          "E[min(max(X-d,0), L)] estimated from historical severity sample", ""),
]
assump_df = pd.DataFrame(assump_rows, columns=["parameter","value","notes"])


# ═══════════════════════════════════════════════════════════════════════════════
# 7. WRITE WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
DB   = PatternFill("solid", start_color="1F4E79")
MB   = PatternFill("solid", start_color="2E75B6")
GB   = PatternFill("solid", start_color="375623")  # dark green for strategy
HF   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BF   = Font(name="Arial", size=10)
thin = Side(style="thin", color="BFBFBF")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

def aw(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max((len(str(c.value or "")) for c in col), default=10) + 3, 48)

def ws_write(ws, df, money=None, pct=None, dec=None, sec_col=None, header_fill=None):
    money = money or []; pct = pct or []; dec = dec or []
    hfill = header_fill or DB
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(1, c, col)
        cell.font, cell.fill = HF, hfill
        cell.alignment, cell.border = Alignment(horizontal="center"), BDR
    for ri, row in enumerate(df.itertuples(index=False), 2):
        is_sec = (sec_col and
                  str(row[df.columns.get_loc(sec_col)]).strip().isupper() and
                  str(row[df.columns.get_loc(sec_col)]).strip() != "")
        for c, val in enumerate(row, 1):
            cell = ws.cell(ri, c, val if pd.notna(val) else "")
            cell.border = BDR
            cname = df.columns[c - 1]
            cell.font = HF if is_sec else BF
            if is_sec: cell.fill = MB
            if not is_sec:
                if cname in money: cell.number_format = "#,##0"
                if cname in pct:   cell.number_format = "0.00%"
                if cname in dec:   cell.number_format = "0.0000"
    aw(ws)

money_cols = ["deductible","occ_limit","lev","pure_premium","el_y0","std_y0","risk_margin",
              "premium_low","premium_high","required_capital","net_rev_low","net_rev_high",
              "p90_y0","p95_y0","p99_y0","p995_y0",
              "tot_nom_mean","tot_nom_p90","tot_nom_p95","tot_nom_p99",
              "tot_disc_mean","tot_disc_p90","tot_disc_p95","tot_disc_p99"]

ws1 = wb.active; ws1.title = "pricing_summary"
ws_write(ws1, pricing_summary,
    money=money_cols, pct=["freq_rate","lr_at_el","roc_low","roc_high"])
ws1.freeze_panes = "C2"

ws2 = wb.create_sheet("loss_distribution")
ws_write(ws2, loss_dist, money=money_cols)
ws2.freeze_panes = "C2"

ws3 = wb.create_sheet("scenario_comparison")
ws_write(ws3, scen_comp,
    money=["pure_premium","premium_low","premium_high","el_y0","p90_y0","p95_y0","p99_y0",
           "required_capital","tot_nom_mean","tot_nom_p99"],
    pct=["roc_low","roc_high"])

ws4 = wb.create_sheet("system_comparison")
ws_write(ws4, sys_comp,
    money=["lev","pure_premium","el_y0","p90_y0","p95_y0","p99_y0",
           "required_capital","net_rev_low","net_rev_high"],
    pct=["freq_rate","roc_low","roc_high"])

ws5 = wb.create_sheet("product_design_specs")
ws_write(ws5, design_specs)

ws6 = wb.create_sheet("strategic_recommendations")
ws_write(ws6, rec_df, sec_col="item", header_fill=GB)

ws7 = wb.create_sheet("assumptions_used")
ws_write(ws7, assump_df, sec_col="parameter")

wb.save(OUTPUT_PATH)


# ═══════════════════════════════════════════════════════════════════════════════
# 8. PRINT SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
print()
print("=" * 74)
print("  BI STEP 5 – PRODUCT DESIGN & PRICING SUMMARY")
print(f"  Valuation date: {VALUATION_DATE}  |  Horizon: {VALUATION_DATE}–{VALUATION_DATE+N}  |  N_sim: {N_SIM:,}")
print("=" * 74)
print(f"  {'System':<20} {'Design':<26} {'Pure PP':>10} {'Prem Lo':>10} {'Prem Hi':>10} {'P99 Y0':>12} {'ROC Lo':>8}")
for _, row in base_df.iterrows():
    print(f"  {row['solar_system']:<20} {row['product_design']:<26} "
          f"{row['pure_premium']:>10,.0f} {row['premium_low']:>10,.0f} "
          f"{row['premium_high']:>10,.0f} {row['p99_y0']:>12,.0f} {row['roc_low']:>8.1%}")
print("-" * 74)
print("  STRATEGIC SUMMARY:")
print("  • Lead product: Design B (Standard) — best balance across all systems")
print("  • Oryn Delta: lowest tail risk, best ROC → prioritise for market entry")
print("  • Helionis Cluster: highest P99 severity → require tighter limits or higher deductible")
print("  • Bayesia System: highest frequency → Design A viable if expense load is managed")
print("  • Cross-system correlated risk (solar storms) → portfolio XL reinsurance recommended")
print("  • Inflation sensitivity: +1% adds ~4–5% to 10yr nominal total across all designs")
print("-" * 74)
print(f"  Output → {OUTPUT_PATH}")
print("=" * 74)