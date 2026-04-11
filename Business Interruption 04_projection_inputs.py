"""
BI Prospective Modelling — Step 1 (Exposure-Led Rebuild)
========================================================
Build a prospective BI exposure dataset from raw Cosmic Quarry files.

PURPOSE
-------
This version is for exposure-based modelling, not proxy-based GLM scoring.

It does NOT construct proxy predictors such as:
  - maintenance_freq
  - energy_backup_score
  - supply_chain_index
  - avg_crew_exp

Instead it builds a clean exposure table with:
  - system
  - equipment type
  - fleet size
  - utilisation
  - active units
  - maintenance hours
  - risk index
  - cargo capacity
  - optional context fields

These fields can support:
  1. exposure-based benchmark modelling
  2. later stress testing
  3. later segmentation if needed

READS:
  /Users/alansteny/Downloads/srcsc-2026-cosmic-quarry-inventory.xlsx
  /Users/alansteny/Downloads/srcsc-2026-cosmic-quarry-personnel.xlsx

WRITES:
  /Users/alansteny/Downloads/bi_prospective_exposure.xlsx
"""

import warnings, textwrap
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INV_PATH = "/Users/alansteny/Downloads/srcsc-2026-cosmic-quarry-inventory.xlsx"
PER_PATH = "/Users/alansteny/Downloads/srcsc-2026-cosmic-quarry-personnel.xlsx"
OUT_PATH = "/Users/alansteny/Downloads/bi_prospective_exposure.xlsx"

SYSTEMS = ["Helionis Cluster", "Bayesia System", "Oryn Delta"]
EQUIP = [
    "Quantum Bores", "Graviton Extractors", "Fexstram Carriers",
    "Regl-Aggregators", "Flux Riders", "Ion Pulverizers"
]

SEP = "=" * 72

print(SEP)
print("STEP 1 — EXPOSURE-LED PROSPECTIVE DATASET")
print(SEP)

# ============================================================================
# A. LOAD RAW FILES
# ============================================================================
print("\nA. LOADING RAW FILES")

raw = pd.read_excel(INV_PATH, header=None)
praw = pd.read_excel(PER_PATH, header=None)

print(f"   Inventory shape : {raw.shape}")
print(f"   Personnel shape : {praw.shape}")

# ============================================================================
# B. PARSE INVENTORY TABLES
# ============================================================================
print("\nB. PARSING INVENTORY TABLES")

# ---- B1. Fleet counts ------------------------------------------------------
fleet = raw.iloc[3:10, 0:4].copy()
fleet.columns = ["equipment_type", "Helionis Cluster", "Bayesia System", "Oryn Delta"]
fleet = fleet.iloc[1:].copy()
fleet["equipment_type"] = (
    fleet["equipment_type"].astype(str).str.strip()
    .replace({"ReglAggregators": "Regl-Aggregators"})
)
for s in SYSTEMS:
    fleet[s] = pd.to_numeric(fleet[s], errors="coerce")
fleet = fleet.reset_index(drop=True)

print(f"   Fleet rows parsed: {len(fleet)}")

# ---- B2. Maintenance / utilisation ----------------------------------------
maint = raw.iloc[39:46, 0:7].copy()
maint.columns = [
    "equipment_type",
    "HC_pct_op", "HC_maint_hrs",
    "BS_pct_op", "BS_maint_hrs",
    "OD_pct_op", "OD_maint_hrs"
]
maint = maint.iloc[1:].copy()
maint["equipment_type"] = (
    maint["equipment_type"].astype(str).str.strip()
    .replace({"ReglAggregators": "Regl-Aggregators"})
)
for c in maint.columns[1:]:
    maint[c] = pd.to_numeric(maint[c], errors="coerce")
maint = maint.reset_index(drop=True)

print(f"   Maintenance rows parsed: {len(maint)}")

# ---- B3. Risk index --------------------------------------------------------
risk = raw.iloc[49:56, 0:4].copy()
risk.columns = ["equipment_type", "Helionis Cluster", "Bayesia System", "Oryn Delta"]
risk = risk.iloc[1:].copy()
risk["equipment_type"] = (
    risk["equipment_type"].astype(str).str.strip()
    .replace({"ReglAggregators": "Regl-Aggregators"})
)
for s in SYSTEMS:
    risk[s] = pd.to_numeric(risk[s], errors="coerce")
risk = risk.reset_index(drop=True)

print(f"   Risk rows parsed: {len(risk)}")

# ---- B4. Fleet age table kept only as raw context --------------------------
AGE_BANDS = {"<5": 2.5, "5-9": 7.0, "10-14": 12.0, "15-19": 17.0, "20+": 22.0}
SYS_AGE_ROWS = {
    "Helionis Cluster": (13, 19),
    "Bayesia System": (21, 27),
    "Oryn Delta": (29, 35),
}

def parse_fleet_age(df_raw, row_start, row_end):
    block = df_raw.iloc[row_start:row_end, :7].copy()
    block.columns = range(block.shape[1])
    acc = {eq: {"w": 0.0, "n": 0.0} for eq in EQUIP}
    for _, row in block.iterrows():
        band = str(row[0]).strip()
        if band not in AGE_BANDS:
            continue
        mid = AGE_BANDS[band]
        for i, eq in enumerate(EQUIP):
            cnt = pd.to_numeric(row[i + 1], errors="coerce")
            if not np.isnan(cnt):
                acc[eq]["w"] += mid * cnt
                acc[eq]["n"] += cnt
    return {eq: (v["w"] / v["n"] if v["n"] > 0 else np.nan) for eq, v in acc.items()}

fleet_age = {
    sys: parse_fleet_age(raw, r0, r1)
    for sys, (r0, r1) in SYS_AGE_ROWS.items()
}

print("   Fleet age parsed as raw context only")

# ---- B5. Cargo capacity ----------------------------------------------------
cargo = raw.iloc[60:66, 0:5].copy()
cargo.columns = ["vessel_type", "max_vol_kg", "Helionis Cluster", "Bayesia System", "Oryn Delta"]
cargo = cargo.iloc[1:].copy()
cargo["max_vol_kg"] = pd.to_numeric(cargo["max_vol_kg"], errors="coerce")
for s in SYSTEMS:
    cargo[s] = pd.to_numeric(cargo[s], errors="coerce")
cargo = cargo.reset_index(drop=True)

sys_cargo_cap = {}
sys_cargo_vessels = {}
for s in SYSTEMS:
    sys_cargo_cap[s] = float((cargo["max_vol_kg"] * cargo[s]).sum())
    sys_cargo_vessels[s] = float(cargo[s].sum())

print("   Cargo capacity parsed")

# ============================================================================
# C. PARSE PERSONNEL AS CONTEXT ONLY
# ============================================================================
print("\nC. PARSING PERSONNEL AS CONTEXT")

praw.columns = range(praw.shape[1])

roles_raw = []
current_dept = None
for _, row in praw.iterrows():
    role = str(row[0]).strip()
    n_emp = pd.to_numeric(row[1], errors="coerce") if row[1] is not None else np.nan
    if np.isnan(n_emp):
        current_dept = role if role not in ("", "nan", "None") else current_dept
        continue
    roles_raw.append({
        "dept": current_dept,
        "role": role,
        "n_emp": n_emp,
        "ft_emp": pd.to_numeric(row[2], errors="coerce"),
        "contract": pd.to_numeric(row[3], errors="coerce"),
        "avg_salary": pd.to_numeric(row[4], errors="coerce"),
        "avg_age": pd.to_numeric(row[5], errors="coerce"),
    })

pers = pd.DataFrame(roles_raw)

ext_pers = pers[pers["dept"] == "Extraction Operations"].copy()
ext_total = float(ext_pers["n_emp"].sum()) if len(ext_pers) else np.nan
ext_ft_rate = float(ext_pers["ft_emp"].sum() / ext_total) if ext_total > 0 else np.nan
ext_avg_age = float((ext_pers["n_emp"] * ext_pers["avg_age"]).sum() / ext_total) if ext_total > 0 else np.nan

personnel_context = pd.DataFrame([
    ["total_personnel", float(pers["n_emp"].sum()), "Company-wide count"],
    ["extraction_personnel", ext_total, "Company-wide extraction operations count"],
    ["extraction_ft_rate", ext_ft_rate, "Company-wide extraction FT rate"],
    ["extraction_avg_age", ext_avg_age, "Company-wide extraction weighted average age"],
], columns=["metric", "value", "note"])

print(f"   Personnel roles parsed: {len(pers)}")
print("   Personnel kept as context only, not row-level predictors")

# ============================================================================
# D. BUILD EXPOSURE TABLE
# ============================================================================
print("\nD. BUILDING EXPOSURE TABLE")

sys_col_map = {
    "Helionis Cluster": {"pct": "HC_pct_op", "mhrs": "HC_maint_hrs"},
    "Bayesia System": {"pct": "BS_pct_op", "mhrs": "BS_maint_hrs"},
    "Oryn Delta": {"pct": "OD_pct_op", "mhrs": "OD_maint_hrs"},
}

records = []
for s in SYSTEMS:
    for eq in EQUIP:
        fleet_row = fleet[fleet["equipment_type"] == eq]
        maint_row = maint[maint["equipment_type"] == eq]
        risk_row = risk[risk["equipment_type"] == eq]

        n_units = float(fleet_row[s].values[0]) if len(fleet_row) else np.nan
        pct_op = float(maint_row[sys_col_map[s]["pct"]].values[0]) if len(maint_row) else np.nan
        maint_hrs = float(maint_row[sys_col_map[s]["mhrs"]].values[0]) if len(maint_row) else np.nan
        risk_idx = float(risk_row[s].values[0]) if len(risk_row) else np.nan
        fleet_avg_svc_yrs = fleet_age.get(s, {}).get(eq, np.nan)

        active_units = n_units * pct_op if not (np.isnan(n_units) or np.isnan(pct_op)) else np.nan

        records.append({
            "solar_system": s,
            "equipment_type": eq,
            "n_units_total": n_units,
            "pct_in_operation": pct_op,
            "active_units": active_units,
            "exposure": active_units,              # main modelling quantity
            "log_exposure": np.log(active_units) if pd.notna(active_units) and active_units > 0 else np.nan,
            "maint_schedule_hrs": maint_hrs,
            "risk_index_raw": risk_idx,
            "fleet_avg_svc_yrs": fleet_avg_svc_yrs,
            "sys_cargo_capacity_kg": sys_cargo_cap[s],
            "sys_cargo_n_vessels": sys_cargo_vessels[s],
        })

df = pd.DataFrame(records)

# exposure shares
total_exposure = df["exposure"].sum()
df["system_exposure_share"] = df.groupby("solar_system")["exposure"].transform(lambda x: x.sum() / total_exposure)
df["equipment_exposure_share"] = df.groupby("equipment_type")["exposure"].transform(lambda x: x.sum() / total_exposure)
df["row_exposure_share"] = df["exposure"] / total_exposure

# optional reference fields only
df["maintenance_hours_per_active_unit"] = df["maint_schedule_hrs"] / df["active_units"]
df["cargo_capacity_per_active_unit"] = df["sys_cargo_capacity_kg"] / df.groupby("solar_system")["active_units"].transform("sum")

# quality / modelling guidance
df["exposure_model_ready"] = 1
df["proxy_model_ready"] = 0
df["modelling_note"] = (
    "Use for exposure-based modelling. Raw exposure is defensible. "
    "Do not treat raw maintenance or risk fields as direct historical GLM equivalents without reconciliation."
)

print(f"   Exposure table rows: {len(df)}")
print(f"   Total exposure     : {total_exposure:.1f}")
print(f"   Exposure range     : [{df['exposure'].min():.1f}, {df['exposure'].max():.1f}]")

# ============================================================================
# E. VALIDATION
# ============================================================================
print("\nE. VALIDATION")

checks_ok = True

req_cols = [
    "solar_system", "equipment_type", "n_units_total", "pct_in_operation",
    "active_units", "exposure", "log_exposure"
]

for col in req_cols:
    if col not in df.columns:
        print(f"   FAIL  missing: {col}")
        checks_ok = False
    elif df[col].isna().any():
        print(f"   WARN  {col} has {int(df[col].isna().sum())} missing")
        checks_ok = False
    else:
        if pd.api.types.is_numeric_dtype(df[col]):
            print(f"   OK    {col:<28} [{df[col].min():.4f}, {df[col].max():.4f}]")
        else:
            print(f"   OK    {col:<28} categorical")

if (df["exposure"] <= 0).any():
    print("   FAIL  exposure must be strictly positive")
    checks_ok = False
else:
    print("   OK    exposure strictly positive")

if len(df) != 18:
    print(f"   FAIL  expected 18 rows, got {len(df)}")
    checks_ok = False
else:
    print("   OK    row count = 18")

# ============================================================================
# F. SUMMARY TABLES
# ============================================================================
print("\nF. BUILDING SUMMARY TABLES")

system_summary = (
    df.groupby("solar_system", as_index=False)
      .agg(
          n_rows=("solar_system", "size"),
          total_units=("n_units_total", "sum"),
          total_active_units=("active_units", "sum"),
          exposure=("exposure", "sum"),
          exposure_share=("row_exposure_share", "sum"),
          avg_pct_in_operation=("pct_in_operation", "mean"),
          avg_maint_schedule_hrs=("maint_schedule_hrs", "mean"),
          avg_risk_index=("risk_index_raw", "mean"),
      )
)

equipment_summary = (
    df.groupby("equipment_type", as_index=False)
      .agg(
          n_systems=("solar_system", "nunique"),
          total_units=("n_units_total", "sum"),
          total_active_units=("active_units", "sum"),
          exposure=("exposure", "sum"),
          exposure_share=("row_exposure_share", "sum"),
          avg_pct_in_operation=("pct_in_operation", "mean"),
          avg_maint_schedule_hrs=("maint_schedule_hrs", "mean"),
          avg_risk_index=("risk_index_raw", "mean"),
      )
      .sort_values("exposure", ascending=False)
      .reset_index(drop=True)
)

# ============================================================================
# G. DATA DICTIONARY
# ============================================================================
print("\nG. BUILDING DATA DICTIONARY")

dd_rows = [
    ("solar_system", "Identifier", "Operating system"),
    ("equipment_type", "Identifier", "Equipment class"),
    ("n_units_total", "Raw", "Total installed units"),
    ("pct_in_operation", "Raw", "Fraction of fleet in operation"),
    ("active_units", "Derived", "n_units_total × pct_in_operation"),
    ("exposure", "Model input", "Primary exposure quantity for exposure-led modelling"),
    ("log_exposure", "Transform", "log(exposure)"),
    ("maint_schedule_hrs", "Raw", "Planned maintenance hours per unit"),
    ("risk_index_raw", "Raw", "Inventory risk index"),
    ("fleet_avg_svc_yrs", "Raw context", "Fleet age only, not crew experience"),
    ("sys_cargo_capacity_kg", "Raw context", "System-level cargo capacity"),
    ("sys_cargo_n_vessels", "Raw context", "System-level vessel count"),
    ("system_exposure_share", "Derived", "System share of total exposure"),
    ("equipment_exposure_share", "Derived", "Equipment share of total exposure"),
    ("row_exposure_share", "Derived", "Row share of total exposure"),
    ("maintenance_hours_per_active_unit", "Reference", "Maintenance intensity per active unit"),
    ("cargo_capacity_per_active_unit", "Reference", "System cargo capacity per active unit"),
    ("exposure_model_ready", "Flag", "1 = ready for exposure-based modelling"),
    ("proxy_model_ready", "Flag", "0 = not intended for proxy GLM scoring"),
    ("modelling_note", "Flag", "Usage guidance"),
]
dd = pd.DataFrame(dd_rows, columns=["column", "category", "description"])

# ============================================================================
# H. EXPORT
# ============================================================================
print("\nH. EXPORTING")

export_df = df.sort_values(["solar_system", "equipment_type"]).reset_index(drop=True)

with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
    export_df.to_excel(writer, sheet_name="Exposure_Dataset", index=False)
    system_summary.to_excel(writer, sheet_name="System_Summary", index=False)
    equipment_summary.to_excel(writer, sheet_name="Equipment_Summary", index=False)
    personnel_context.to_excel(writer, sheet_name="Personnel_Context", index=False)
    dd.to_excel(writer, sheet_name="Data_Dictionary", index=False)

    wb = writer.book
    H_FILL = PatternFill("solid", fgColor="1B2A4A")
    GREY = PatternFill("solid", fgColor="F4F6F6")
    GREEN = PatternFill("solid", fgColor="D5F5E3")
    BLUE = PatternFill("solid", fgColor="EAF2F8")
    WHITE = PatternFill("solid", fgColor="FDFEFE")
    thin = Side(style="thin", color="BDC3C7")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sh_name in wb.sheetnames:
        ws = wb[sh_name]

        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = H_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
        ws.row_dimensions[1].height = 30

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="center")
                cell.border = bdr

                if sh_name == "Exposure_Dataset":
                    hdr = ws.cell(1, cell.column).value
                    if hdr in ["solar_system", "equipment_type"]:
                        cell.fill = BLUE
                    elif hdr in ["exposure", "log_exposure", "system_exposure_share", "equipment_exposure_share", "row_exposure_share"]:
                        cell.fill = GREEN
                    else:
                        cell.fill = WHITE
                else:
                    cell.fill = GREY

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)

        ws.freeze_panes = "A2"

print(f"   Export complete: {OUT_PATH}")

# ============================================================================
# I. FINAL SUMMARY
# ============================================================================
print(f"\n{SEP}")
print("STEP 1 — EXPOSURE-LED SUMMARY")
print(SEP)
print("This version is built for exposure-based modelling.")
print("It keeps raw operational quantities and a defensible exposure measure.")
print("It does not create proxy GLM predictors for scoring.")
print("")
print("Primary modelling field:")
print("  exposure = active_units = n_units_total × pct_in_operation")
print("")
print("Use cases:")
print("  1. exposure-based benchmark modelling")
print("  2. relative portfolio share analysis")
print("  3. later stress testing and scenario design")
print("")
print("Not recommended from this file without further reconciliation:")
print("  1. direct GLM scoring with historical proxy coefficients")
print("  2. absolute pricing using mismatched proxy scales")
print("")
print("Preview:")
print(export_df[[
    "solar_system", "equipment_type", "n_units_total",
    "pct_in_operation", "exposure", "row_exposure_share"
]].to_string(index=False))
print(f"\n{SEP}")
print("Exposure-led Step 1 complete.")
print(SEP)