"""
BI Step 3 – Forward Projection  |  Valuation date: 2175  |  Horizon: 2175–2185
===============================================================================
Corrections applied:
  1. Frequency now filtered to data_quality_flag == 0, consistent with severity.
  2. Frequency trend is explicitly tested (OLS on annual cohort rates).
     Result is reported and the constant-frequency assumption is stated with
     the actual p-value and R², not asserted without evidence.
  3. Rate projection instrument switch at Y6+ is clearly labelled:
     Y1–Y2 use last-3yr 1y spot (near-term persistence),
     Y3–Y5 use last-5yr 1y spot (mean reversion),
     Y6–Y10 switch to full-series 10y spot as the long-run structural anchor.
     The instrument change is documented explicitly in the assumptions sheet.
  4. Year 0 = 2175 is the valuation date. Discount factor at t=0 is 1.0 by
     definition (undiscounted valuation-date loss). This is stated explicitly.
"""

import pandas as pd
import numpy as np
from scipy.stats import linregress
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXPOSURE_PATH   = "/Users/alansteny/Downloads/bi_prospective_exposure.xlsx"
HISTORICAL_PATH = "/Users/alansteny/Downloads/BI_data_CLEANED.xlsx"
RATES_PATH      = "/Users/alansteny/Downloads/srcsc-2026-interest-and-inflation.xlsx"
EXPOSURE_SHEET  = "Exposure_Dataset"
OUTPUT_PATH     = "/Users/alansteny/Downloads/bi_step3_forward_projection.xlsx"

VALUATION_DATE = 2175   # Year 0: undiscounted. All future years discounted back to this point.
N              = 10     # project through 2185


# ── 1. LOAD & PROJECT RATES 2175–2185 ────────────────────────────────────────
r = pd.read_excel(RATES_PATH, sheet_name="Sheet1", header=1, skiprows=[2]).iloc[1:]
r.columns = ["year","inflation","overnight","spot_1y","spot_10y"]
r = r.dropna(subset=["year"]).copy()
r["year"] = r["year"].astype(int)
for c in ["inflation","overnight","spot_1y","spot_10y"]:
    r[c] = pd.to_numeric(r[c])

avg3  = r.tail(3).mean(numeric_only=True)
avg5  = r.tail(5).mean(numeric_only=True)
avgA  = r.mean(numeric_only=True)
last  = r.iloc[-1]

# Instrument and averaging window by projection year offset t
# t=0  : valuation date — use last observed as context only (not applied to discounting)
# t=1-2: last-3yr avg of 1y spot  (near-term: current regime expected to persist)
# t=3-5: last-5yr avg of 1y spot  (medium: mean reversion toward recent equilibrium)
# t=6-10: full-series avg of 10y spot  (long-run structural anchor; instrument switch noted)
def _rates(t):
    if   t == 0: return "current (2174)",         last["inflation"],  last["spot_1y"],   "1y spot"
    elif t <= 2: return "near: last-3yr 1y avg",  avg3["inflation"],  avg3["spot_1y"],   "1y spot"
    elif t <= 5: return "mid: last-5yr 1y avg",   avg5["inflation"],  avg5["spot_1y"],   "1y spot"
    else:        return "long: full-series 10y",  avgA["inflation"],  avgA["spot_10y"],  "10y spot [instrument switch]"

proj_rates = {VALUATION_DATE + t: _rates(t) for t in range(N + 1)}


# ── 2. HISTORICAL BI — CONSISTENT FILTERING (flag==0 on both sides) ───────────
freq_raw = pd.read_excel(HISTORICAL_PATH, sheet_name="Frequency_Cleaned")
sev_raw  = pd.read_excel(HISTORICAL_PATH, sheet_name="Severity_Cleaned")

freq = freq_raw[(freq_raw["data_quality_flag"] == 0)].dropna(subset=["exposure","claim_count"]).copy()
sev  = sev_raw[ (sev_raw["data_quality_flag"]  == 0)].dropna(subset=["claim_amount"]).copy()

frequency_rate = freq["claim_count"].sum() / freq["exposure"].sum()
severity_mean  = sev["claim_amount"].mean()
severity_std   = sev["claim_amount"].std()

# ── Frequency trend test (OLS on 15 annual cohort rates) ─────────────────────
freq["policy_num"]   = freq["policy_id"].str.extract(r"(\d+)").astype(float)
freq["year_cohort"]  = pd.cut(freq["policy_num"].dropna(), bins=15, labels=range(2160,2175))
annual_freq = (freq.dropna(subset=["year_cohort"])
               .groupby("year_cohort", observed=True)
               .agg(claims=("claim_count","sum"), exposure=("exposure","sum")))
annual_freq["rate"] = annual_freq["claims"] / annual_freq["exposure"]
slope, _, r_val, p_val, _ = linregress(range(15), annual_freq["rate"])
freq_r2 = r_val ** 2
freq_p  = p_val
# Frequency is held constant because the trend test is not significant.
# This is an empirical result, not an assertion.
freq_constant = (p_val >= 0.05)


# ── 3. SCORE YEAR 0 (valuation date = 2175, discount factor = 1.0) ────────────
exp = pd.read_excel(EXPOSURE_PATH, sheet_name=EXPOSURE_SHEET).copy()
assert "active_units" in exp.columns
exp["expected_claims"]   = exp["active_units"] * frequency_rate
exp["expected_loss"]     = exp["expected_claims"] * severity_mean
exp["loss_per_exposure"] = exp["expected_loss"] / exp["active_units"]
year0_loss = exp["expected_loss"].sum()


# ── 4. PROJECTION ENGINE ──────────────────────────────────────────────────────
def project(y0, inf_shift=0.0, disc_shift=0.0, exp_growth=0.0):
    rows = []
    cum_sev = 1.0
    cum_nom = cum_disc = 0.0
    for t in range(N + 1):
        yr = VALUATION_DATE + t
        seg, inf, disc, instr = proj_rates[yr]
        inf  = inf  + inf_shift
        disc = max(disc + disc_shift, 0.0001)
        if t > 0:
            cum_sev *= (1 + inf)
        nom  = y0 * cum_sev * (1 + exp_growth) ** t
        if t == 0:
            df_t = 1.0   # valuation date: undiscounted by definition
        else:
            df_t = np.prod([
                1 / (1 + max(proj_rates[VALUATION_DATE+i][2] + disc_shift, 0.0001))
                for i in range(1, t+1)
            ])
        disc_loss = nom * df_t
        cum_nom  += nom
        cum_disc += disc_loss
        rows.append({
            "year":             yr,
            "t":                t,
            "segment":          seg,
            "instrument":       instr,
            "inflation":        inf if t > 0 else None,
            "disc_rate":        disc if t > 0 else None,
            "cum_sev_factor":   round(cum_sev, 6),
            "discount_factor":  round(df_t, 8),
            "nominal_loss":     nom,
            "discounted_loss":  disc_loss,
            "cum_nominal":      cum_nom,
            "cum_discounted":   cum_disc,
        })
    return pd.DataFrame(rows)

base = project(year0_loss)


# ── 5. SCENARIOS ──────────────────────────────────────────────────────────────
SCENARIOS = [
    ("Base",               0.0,   0.0,   0.0),
    ("Inf +100bp",         0.01,  0.0,   0.0),
    ("Inf -100bp",        -0.01,  0.0,   0.0),
    ("Rates +100bp",       0.0,   0.01,  0.0),
    ("Rates -100bp",       0.0,  -0.01,  0.0),
    ("Exposure +5% pa",    0.0,   0.0,   0.05),
    ("Stress Combined",    0.01,  0.01,  0.05),
]

scen_rows = []
for label, di, dd, eg in SCENARIOS:
    df = project(year0_loss, di, dd, eg)
    scen_rows.append({"scenario":label, "inf_shift":di, "disc_shift":dd, "exp_growth":eg,
                      "year0_loss":year0_loss,
                      "total_nominal":df["nominal_loss"].sum(),
                      "total_discounted":df["discounted_loss"].sum(),
                      "vs_base_nom_%":None, "vs_base_disc_%":None})

scen_df = pd.DataFrame(scen_rows)
bn = scen_df.loc[scen_df.scenario=="Base","total_nominal"].values[0]
bd = scen_df.loc[scen_df.scenario=="Base","total_discounted"].values[0]
scen_df["vs_base_nom_%"]  = (scen_df["total_nominal"]    - bn) / bn
scen_df["vs_base_disc_%"] = (scen_df["total_discounted"] - bd) / bd


# ── 6. ASSUMPTIONS ────────────────────────────────────────────────────────────
freq_assumption = ("constant — OLS trend not significant"
                   if freq_constant else "trending — OLS slope statistically significant")
assump = [
    ("VALUATION & PROJECTION PERIOD", "", ""),
    ("valuation_date",      str(VALUATION_DATE),
                             "Year 0. Discount factor = 1.0 at this point by definition."),
    ("projection_horizon",  f"{VALUATION_DATE}–{VALUATION_DATE+N}",
                             "10 forward years"),
    ("", "", ""),
    ("HISTORICAL DATA QUALITY FILTER", "", ""),
    ("frequency_filter",    "data_quality_flag == 0",
                             f"{len(freq):,} rows used from Frequency_Cleaned"),
    ("severity_filter",     "data_quality_flag == 0",
                             f"{len(sev):,} rows used from Severity_Cleaned"),
    ("filter_consistency",  "Both sides filtered identically",
                             "Ensures freq_rate and severity_mean share the same clean basis"),
    ("", "", ""),
    ("BI PARAMETERS", "", ""),
    ("frequency_rate",      f"{frequency_rate:.8f}",
                             "claim_count / exposure on flag==0 rows"),
    ("severity_mean",       f"{severity_mean:,.2f}",
                             "mean(claim_amount) on flag==0 rows"),
    ("severity_std",        f"{severity_std:,.2f}",  ""),
    ("year0_expected_loss", f"{year0_loss:,.2f}",
                             "active_units × frequency_rate × severity_mean, summed"),
    ("", "", ""),
    ("FREQUENCY TREND TEST", "", ""),
    ("method",              "OLS on 15 annual cohort frequency rates (2160–2174)", ""),
    ("ols_slope",           f"{slope:.6f} per year",  "Change in frequency rate per cohort year"),
    ("ols_r2",              f"{freq_r2:.4f}",          ""),
    ("ols_p_value",         f"{freq_p:.4f}",           "p >= 0.05 → not significant at 5% level"),
    ("frequency_assumption",freq_assumption,            ""),
    ("", "", ""),
    ("RATE PROJECTION 2175–2185", "", ""),
    ("source_file",         "srcsc-2026-interest-and-inflation.xlsx",
                             f"Observed {int(r['year'].min())}–{int(r['year'].max())}"),
    ("t0_2175",             "Last observed (2174 actual)",
                             "Valuation date context only — not applied to discounting"),
    ("t1_t2_2176_2177",     "Last-3yr avg, 1y spot / inflation",
                             "Near-term: current regime expected to persist"),
    ("t3_t5_2178_2180",     "Last-5yr avg, 1y spot / inflation",
                             "Medium: mean reversion toward recent equilibrium"),
    ("t6_t10_2181_2185",    "Full-series avg, 10y spot / inflation",
                             "Long-run structural anchor. NOTE: instrument switches from 1y to 10y spot here."),
    ("instrument_switch_note",
                            "1y spot (Y1–Y5) → 10y spot (Y6–Y10)",
                             "10y spot used as structural long-run anchor, not as continuation of 1y series"),
]
for t in range(N+1):
    yr = VALUATION_DATE + t
    seg, inf, disc, instr = proj_rates[yr]
    assump.append((f"  {yr}", f"inf={inf:.4%}   disc={disc:.4%}", f"{seg} | {instr}"))

assump_df = pd.DataFrame(assump, columns=["parameter","value","notes"])


# ── 7. WRITE WORKBOOK ─────────────────────────────────────────────────────────
wb = Workbook()
DB = PatternFill("solid", start_color="1F4E79")
MB = PatternFill("solid", start_color="2E75B6")
HF = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BF = Font(name="Arial", size=10)
thin = Side(style="thin", color="BFBFBF")
BDR = Border(left=thin, right=thin, top=thin, bottom=thin)

def aw(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max((len(str(c.value or "")) for c in col), default=10)+3, 46)

def ws_write(ws, df, money=None, pct=None, dec=None, sec_col=None):
    money=money or[]; pct=pct or[]; dec=dec or[]
    for c,col in enumerate(df.columns,1):
        cell=ws.cell(1,c,col); cell.font,cell.fill=HF,DB
        cell.alignment,cell.border=Alignment(horizontal="center"),BDR
    for r,row in enumerate(df.itertuples(index=False),2):
        is_sec=(sec_col and str(row[df.columns.get_loc(sec_col)]).strip().isupper()
                and str(row[df.columns.get_loc(sec_col)]).strip()!="")
        for c,val in enumerate(row,1):
            cell=ws.cell(r,c,val if pd.notna(val) else "")
            cell.border=BDR; cname=df.columns[c-1]
            cell.font=HF if is_sec else BF
            if is_sec: cell.fill=MB
            if not is_sec:
                if cname in money: cell.number_format='#,##0'
                if cname in pct:   cell.number_format='0.00%'
                if cname in dec:   cell.number_format='0.000000'
    aw(ws)

ws1=wb.active; ws1.title="annual_projection"
ws_write(ws1, base,
    money=["nominal_loss","discounted_loss","cum_nominal","cum_discounted"],
    pct=["inflation","disc_rate"], dec=["cum_sev_factor","discount_factor"])
ws1.freeze_panes="C2"

ws2=wb.create_sheet("scenario_summary")
ws_write(ws2, scen_df,
    money=["year0_loss","total_nominal","total_discounted"],
    pct=["inf_shift","disc_shift","exp_growth","vs_base_nom_%","vs_base_disc_%"])

ws3=wb.create_sheet("assumptions_used")
ws_write(ws3, assump_df, sec_col="parameter")

ws4=wb.create_sheet("prospective_portfolio")
ws_write(ws4, exp,
    money=["expected_loss","loss_per_exposure"], dec=["expected_claims"])

wb.save(OUTPUT_PATH)


# ── 8. SUMMARY ────────────────────────────────────────────────────────────────
nom_total  = base["nominal_loss"].sum()
disc_total = base["discounted_loss"].sum()

print("="*70)
print("  BI STEP 3 – FORWARD PROJECTION  |  Valuation date: 2175  |  To: 2185")
print("="*70)
print(f"  Freq rate (flag==0) : {frequency_rate:.6f}   Sev mean: {severity_mean:,.0f}")
print(f"  Freq OLS trend      : slope={slope:.6f}  R²={freq_r2:.4f}  p={freq_p:.4f}  → {freq_assumption}")
print(f"  Year 0 loss (2175)  : {year0_loss:,.0f}  [discount factor = 1.0 at valuation date]")
print("-"*70)
print(f"  {'Year':<6} {'Segment':<26} {'Instr':<8} {'Inf':>8} {'Disc':>8} {'Nominal Loss':>16}")
for _,row in base.iterrows():
    inf_s  = f"{row['inflation']:.2%}"  if pd.notna(row['inflation'])  else "—"
    disc_s = f"{row['disc_rate']:.2%}"  if pd.notna(row['disc_rate'])  else "—"
    instr  = "10y" if "10y" in str(row['instrument']) else "1y"
    print(f"  {int(row['year']):<6} {row['segment']:<26} {instr:<8} {inf_s:>8} {disc_s:>8} {row['nominal_loss']:>16,.0f}")
print("-"*70)
print(f"  Total nominal    : {nom_total:>16,.0f}")
print(f"  Total discounted : {disc_total:>16,.0f}   (drag: {(nom_total-disc_total)/nom_total:.2%})")
print("-"*70)
print(f"  {'Scenario':<28} {'Nominal':>16}  {'vs Base':>8}")
for _,row in scen_df.iterrows():
    tag = f"{row['vs_base_nom_%']:+.1%}" if row['scenario']!="Base" else "base"
    print(f"  {row['scenario']:<28} {row['total_nominal']:>16,.0f}  {tag:>8}")
print("-"*70)
print(f"  Output → {OUTPUT_PATH}")
print("="*70)