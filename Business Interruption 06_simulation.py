"""
BI Step 4 – Stochastic Loss Simulation  |  2175–2185
=====================================================
Extends the Step 3 deterministic baseline by introducing distributional
uncertainty in claim counts and claim sizes.

Frequency : Negative Binomial — overdispersion ratio var/mean = 1.717
            (Poisson would force ratio = 1.0, understating volatility).
            NB parameterised as: r_t = lambda_t / (od - 1), p = 1 / od
            where lambda_t = frequency_rate × exposure_t.

Severity  : Lognormal fitted to cleaned historical claims (flag==0, amount>0).
            KS test p = 0.752 — not rejected at 5%.

Exposure  : Three scenarios — constant (base), +5% pa (growth), -2% pa (contraction).
            Expected claims in year t = frequency_rate × exposure_0 × (1+g)^t.

Inflation / discounting : inherited directly from Step 3 term structure.
Valuation date : 2175 (t=0, discount factor = 1.0).
"""

import pandas as pd
import numpy as np
from scipy import stats
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PATHS ─────────────────────────────────────────────────────────────────────
EXPOSURE_PATH   = "/Users/alansteny/Downloads/bi_prospective_exposure.xlsx"
HISTORICAL_PATH = "/Users/alansteny/Downloads/BI_data_CLEANED.xlsx"
RATES_PATH      = "/Users/alansteny/Downloads/srcsc-2026-interest-and-inflation.xlsx"
EXPOSURE_SHEET  = "Exposure_Dataset"
OUTPUT_PATH     = "/Users/alansteny/Downloads/bi_step4_simulation.xlsx"

VALUATION_DATE = 2175
N              = 10        # years: 2175–2185
N_SIM          = 10_000
RNG_SEED       = 42

# Exposure growth scenarios: label → annual growth rate
EXPOSURE_SCENARIOS = {
    "Base (constant)":    0.00,
    "Growth (+5% pa)":    0.05,
    "Contraction (-2% pa)": -0.02,
}

PCTS = [50, 75, 90, 95, 99]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. RATES — Step 3 term structure (replicated exactly)
# ═══════════════════════════════════════════════════════════════════════════════
r = pd.read_excel(RATES_PATH, sheet_name="Sheet1", header=1, skiprows=[2]).iloc[1:]
r.columns = ["year", "inflation", "overnight", "spot_1y", "spot_10y"]
r = r.dropna(subset=["year"]).copy()
r["year"] = r["year"].astype(int)
for c in ["inflation", "overnight", "spot_1y", "spot_10y"]:
    r[c] = pd.to_numeric(r[c])

avg3 = r.tail(3).mean(numeric_only=True)
avg5 = r.tail(5).mean(numeric_only=True)
avgA = r.mean(numeric_only=True)
last = r.iloc[-1]

def _rates(t):
    if   t == 0: return last["inflation"], last["spot_1y"],  "current (2174)"
    elif t <= 2: return avg3["inflation"], avg3["spot_1y"],  "near: last-3yr 1y avg"
    elif t <= 5: return avg5["inflation"], avg5["spot_1y"],  "mid: last-5yr 1y avg"
    else:        return avgA["inflation"], avgA["spot_10y"], "long: full-series 10y avg"

# Precompute cumulative factors (same as Step 3)
cum_inf  = {}   # cumulative severity inflation at year t
cum_disc = {}   # cumulative discount factor at year t
ci = cd = 1.0
for t in range(N + 1):
    if t > 0:
        inf_t, disc_t, _ = _rates(t)
        ci *= (1 + inf_t)
        cd *= 1 / (1 + disc_t)
    cum_inf[t]  = ci
    cum_disc[t] = cd if t > 0 else 1.0


# ═══════════════════════════════════════════════════════════════════════════════
# 2. HISTORICAL PARAMETERS — consistent flag==0 filter
# ═══════════════════════════════════════════════════════════════════════════════
freq_raw = pd.read_excel(HISTORICAL_PATH, sheet_name="Frequency_Cleaned")
sev_raw  = pd.read_excel(HISTORICAL_PATH, sheet_name="Severity_Cleaned")

freq = freq_raw[freq_raw["data_quality_flag"] == 0].dropna(subset=["exposure", "claim_count"]).copy()
sev  = sev_raw[(sev_raw["data_quality_flag"] == 0) & (sev_raw["claim_amount"] > 0)
               ].dropna(subset=["claim_amount"]).copy()

frequency_rate = freq["claim_count"].sum() / freq["exposure"].sum()

# ── Negative Binomial overdispersion (method of moments on per-policy counts) ─
cc      = freq["claim_count"].values
mu_cc   = cc.mean()
var_cc  = cc.var(ddof=1)
od      = var_cc / mu_cc          # overdispersion ratio: 1.717
# Portfolio NB parameterisation:
#   lambda_t = freq_rate × exposure_t
#   r_t      = lambda_t / (od - 1)   [varies with exposure scenario and year]
#   p_nb     = 1 / od                 [fixed — encodes dispersion structure]
p_nb    = 1.0 / od

# ── Lognormal severity (MLE closed form) ──────────────────────────────────────
x       = sev["claim_amount"].values
log_x   = np.log(x)
ln_mu   = log_x.mean()
ln_sig  = log_x.std(ddof=1)
ks_stat, ks_p = stats.kstest(x, "lognorm", args=(ln_sig, 0, np.exp(ln_mu)))
sev_n   = len(x)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. PROSPECTIVE EXPOSURE
# ═══════════════════════════════════════════════════════════════════════════════
exp_df = pd.read_excel(EXPOSURE_PATH, sheet_name=EXPOSURE_SHEET)
assert "active_units" in exp_df.columns
exposure_0 = exp_df["active_units"].sum()


# ═══════════════════════════════════════════════════════════════════════════════
# 4. SIMULATION ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
rng = np.random.default_rng(RNG_SEED)

def simulate(growth_rate):
    """
    Run N_SIM iterations for one exposure scenario.
    Returns:
        nominal_matrix    (N_SIM, N+1)
        discounted_matrix (N_SIM, N+1)
        lambda_t_list     list of expected claims per year
    """
    nominal_mat    = np.zeros((N_SIM, N + 1))
    discounted_mat = np.zeros((N_SIM, N + 1))
    lambda_t_list  = []

    for t in range(N + 1):
        exposure_t = exposure_0 * (1 + growth_rate) ** t
        lambda_t   = frequency_rate * exposure_t
        lambda_t_list.append(lambda_t)

        # NB parameters for this year's portfolio-level draw
        r_t = lambda_t / (od - 1)   # r scales with expected claims

        # Draw claim counts: NB(n=r_t, p=p_nb) — shape (N_SIM,)
        counts = rng.negative_binomial(n=r_t, p=p_nb, size=N_SIM)

        # Draw severities vectorised: (N_SIM, max_count) then mask
        max_c = int(counts.max())
        if max_c == 0:
            raw_losses = np.zeros(N_SIM)
        else:
            sev_mat = rng.lognormal(mean=ln_mu, sigma=ln_sig, size=(N_SIM, max_c))
            mask    = np.arange(max_c)[None, :] < counts[:, None]
            raw_losses = (sev_mat * mask).sum(axis=1)

        nominal_mat[:, t]    = raw_losses * cum_inf[t]
        discounted_mat[:, t] = raw_losses * cum_inf[t] * cum_disc[t]

    return nominal_mat, discounted_mat, lambda_t_list


# ── Run all scenarios ──────────────────────────────────────────────────────────
results = {}
for scenario, g in EXPOSURE_SCENARIOS.items():
    print(f"  Simulating: {scenario} ...", flush=True)
    nom, disc, lam = simulate(g)
    results[scenario] = {
        "nominal":    nom,
        "discounted": disc,
        "lambda_t":   lam,
        "growth":     g,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 5. BUILD OUTPUT TABLES
# ═══════════════════════════════════════════════════════════════════════════════
def pct_row(label, arr):
    row = {"scenario": label,
           "mean":     arr.mean(),
           "median":   np.median(arr),
           "std":      arr.std(ddof=1)}
    for p in PCTS:
        row[f"P{p}"] = np.percentile(arr, p)
    return row

# simulation_summary — one row per scenario × metric
summary_rows = []
for scen, res in results.items():
    tot_nom  = res["nominal"].sum(axis=1)
    tot_disc = res["discounted"].sum(axis=1)
    summary_rows.append(pct_row(f"{scen} | Nominal",    tot_nom))
    summary_rows.append(pct_row(f"{scen} | Discounted", tot_disc))
sim_summary = pd.DataFrame(summary_rows)

# scenario_comparison — side-by-side P50/P90/P95/P99 nominal totals
comp_rows = []
for scen, res in results.items():
    tot = res["nominal"].sum(axis=1)
    comp_rows.append({
        "scenario":      scen,
        "exposure_growth": res["growth"],
        "mean":           tot.mean(),
        "P50":            np.percentile(tot, 50),
        "P75":            np.percentile(tot, 75),
        "P90":            np.percentile(tot, 90),
        "P95":            np.percentile(tot, 95),
        "P99":            np.percentile(tot, 99),
        "vs_base_mean_%": None,
    })
comp_df = pd.DataFrame(comp_rows)
base_mean = comp_df.loc[comp_df["scenario"] == "Base (constant)", "mean"].values[0]
comp_df["vs_base_mean_%"] = (comp_df["mean"] - base_mean) / base_mean

# annual_distribution — base scenario only, nominal
base_res = results["Base (constant)"]
annual_rows = []
for t in range(N + 1):
    yr  = VALUATION_DATE + t
    nom = base_res["nominal"][:, t]
    _, _, seg = _rates(t)
    row = {"year": yr, "t": t, "segment": seg,
           "expected_claims": base_res["lambda_t"][t],
           "cum_inf_factor":  cum_inf[t],
           "cum_disc_factor": cum_disc[t],
           "mean":    nom.mean(),
           "median":  np.median(nom),
           "std":     nom.std(ddof=1)}
    for p in PCTS:
        row[f"P{p}"] = np.percentile(nom, p)
    annual_rows.append(row)
annual_dist = pd.DataFrame(annual_rows)

# severity_fit
sev_fit = pd.DataFrame([
    ("distribution",       "Lognormal",                   "Selected: KS p=0.752; Gamma rejected KS p=0.000"),
    ("ln_mu",              f"{ln_mu:.6f}",                "MLE log-mean: mean(log(claim_amount))"),
    ("ln_sigma",           f"{ln_sig:.6f}",               "MLE log-std: std(log(claim_amount), ddof=1)"),
    ("implied_mean",       f"{np.exp(ln_mu+ln_sig**2/2):,.2f}", "exp(mu + sigma²/2)"),
    ("implied_median",     f"{np.exp(ln_mu):,.2f}",        "exp(mu)"),
    ("implied_std",        f"{np.sqrt((np.exp(ln_sig**2)-1)*np.exp(2*ln_mu+ln_sig**2)):,.2f}", ""),
    ("observed_mean",      f"{x.mean():,.2f}",             ""),
    ("observed_median",    f"{np.median(x):,.2f}",         ""),
    ("observed_std",       f"{x.std(ddof=1):,.2f}",        ""),
    ("sample_size",        str(sev_n),                     "flag==0, claim_amount > 0"),
    ("KS_statistic",       f"{ks_stat:.4f}",               ""),
    ("KS_p_value",         f"{ks_p:.4f}",                  "Not rejected at 5% level"),
], columns=["parameter", "value", "notes"])

# frequency_fit
freq_fit = pd.DataFrame([
    ("distribution",        "Negative Binomial",           "Overdispersion ratio=1.717 — Poisson (ratio=1.0) understates variance"),
    ("overdispersion_ratio",f"{od:.6f}",                   "var(claim_count) / mean(claim_count) per policy row"),
    ("p_nb_fixed",          f"{p_nb:.6f}",                 "p = 1 / od — fixed across all years and scenarios"),
    ("r_t_formula",         "lambda_t / (od - 1)",         "r scales with expected claims; computed per year"),
    ("frequency_rate",      f"{frequency_rate:.8f}",       "total claim_count / total exposure, flag==0 rows"),
    ("historical_n_rows",   f"{len(freq):,}",              "Frequency_Cleaned, flag==0"),
    ("mu_per_policy",       f"{mu_cc:.6f}",                "mean claim_count per policy row"),
    ("var_per_policy",      f"{var_cc:.6f}",               "var claim_count per policy row (ddof=1)"),
], columns=["parameter", "value", "notes"])

# assumptions_used
assump = [
    ("SIMULATION SETUP", "", ""),
    ("n_simulations",       str(N_SIM),          ""),
    ("rng_seed",            str(RNG_SEED),        "Reproducible"),
    ("valuation_date",      str(VALUATION_DATE),  "t=0; discount factor = 1.0 by definition"),
    ("horizon",             f"{VALUATION_DATE}–{VALUATION_DATE+N}", ""),
    ("", "", ""),
    ("FREQUENCY", "", ""),
    ("distribution",        "Negative Binomial(r_t, p_nb)", ""),
    ("r_t",                 "lambda_t / (od - 1)",  "Varies by year and exposure scenario"),
    ("p_nb",                f"{p_nb:.6f}",           "Fixed: 1 / overdispersion_ratio"),
    ("overdispersion",      f"{od:.4f}",             "Empirical var/mean from historical policy counts"),
    ("lambda_t",            "freq_rate × exposure_0 × (1+g)^t", ""),
    ("", "", ""),
    ("SEVERITY", "", ""),
    ("distribution",        "Lognormal(ln_mu, ln_sigma)", ""),
    ("ln_mu",               f"{ln_mu:.6f}",          ""),
    ("ln_sigma",            f"{ln_sig:.6f}",          ""),
    ("sample_n",            str(sev_n),               "flag==0, claim_amount > 0"),
    ("", "", ""),
    ("EXPOSURE SCENARIOS", "", ""),
]
for scen, g in EXPOSURE_SCENARIOS.items():
    assump.append((f"  {scen}", f"g={g:.2%}", "exposure_0 × (1+g)^t"))

assump += [
    ("exposure_0",          f"{exposure_0:,.4f}",    "Sum of active_units, Exposure_Dataset"),
    ("", "", ""),
    ("INFLATION & DISCOUNTING (Step 3 term structure)", "", ""),
    ("t1_t2",   f"inf={avg3['inflation']:.4%}  disc={avg3['spot_1y']:.4%}",  "Last-3yr avg, 1y spot"),
    ("t3_t5",   f"inf={avg5['inflation']:.4%}  disc={avg5['spot_1y']:.4%}",  "Last-5yr avg, 1y spot"),
    ("t6_t10",  f"inf={avgA['inflation']:.4%}  disc={avgA['spot_10y']:.4%}", "Full-series avg, 10y spot [instrument switch from 1y]"),
    ("inflation_application", "Cumulative ∏(1+inf_i) for i=1..t applied to aggregate loss", ""),
    ("discounting",           "Cumulative ∏(1/(1+disc_i)) for i=1..t",                      ""),
]
assump_df = pd.DataFrame(assump, columns=["parameter", "value", "notes"])


# ═══════════════════════════════════════════════════════════════════════════════
# 6. WRITE WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
DB   = PatternFill("solid", start_color="1F4E79")
MB   = PatternFill("solid", start_color="2E75B6")
HF   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BF   = Font(name="Arial", size=10)
thin = Side(style="thin", color="BFBFBF")
BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)

def aw(ws):
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max((len(str(c.value or "")) for c in col), default=10) + 3, 46)

def ws_write(ws, df, money=None, pct=None, dec=None, sec_col=None):
    money = money or []; pct = pct or []; dec = dec or []
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(1, c, col)
        cell.font, cell.fill = HF, DB
        cell.alignment, cell.border = Alignment(horizontal="center"), BDR
    for ri, row in enumerate(df.itertuples(index=False), 2):
        is_sec = (sec_col and
                  str(row[df.columns.get_loc(sec_col)]).strip().isupper() and
                  str(row[df.columns.get_loc(sec_col)]).strip() != "")
        for c, val in enumerate(row, 1):
            cell = ws.cell(ri, c, val if pd.notna(val) else "")
            cell.border = BDR
            cname = df.columns[c - 1]
            cell.font = HF if is_sec else BF
            if is_sec: cell.fill = MB
            if not is_sec:
                if cname in money: cell.number_format = "#,##0"
                if cname in pct:   cell.number_format = "0.00%"
                if cname in dec:   cell.number_format = "0.000000"
    aw(ws)

money_stat = ["mean", "median", "std"] + [f"P{p}" for p in PCTS]

ws1 = wb.active; ws1.title = "simulation_summary"
ws_write(ws1, sim_summary, money=money_stat)

ws2 = wb.create_sheet("scenario_comparison")
ws_write(ws2, comp_df,
    money=["mean","P50","P75","P90","P95","P99"],
    pct=["exposure_growth","vs_base_mean_%"])

ws3 = wb.create_sheet("annual_distribution")
ws_write(ws3, annual_dist,
    money=["mean","median","std"] + [f"P{p}" for p in PCTS],
    dec=["cum_inf_factor","cum_disc_factor","expected_claims"])
ws3.freeze_panes = "C2"

ws4 = wb.create_sheet("severity_fit")
ws_write(ws4, sev_fit)

ws5 = wb.create_sheet("frequency_fit")
ws_write(ws5, freq_fit)

ws6 = wb.create_sheet("assumptions_used")
ws_write(ws6, assump_df, sec_col="parameter")

wb.save(OUTPUT_PATH)


# ═══════════════════════════════════════════════════════════════════════════════
# 7. PRINT SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("  BI STEP 4 – STOCHASTIC SIMULATION SUMMARY")
print(f"  Valuation date: {VALUATION_DATE}  |  Horizon: {VALUATION_DATE}–{VALUATION_DATE+N}  |  Iterations: {N_SIM:,}")
print("=" * 70)
print(f"  Frequency : Negative Binomial  od={od:.4f}  p={p_nb:.6f}")
print(f"  Severity  : Lognormal  mu={ln_mu:.4f}  sigma={ln_sig:.4f}  KS p={ks_p:.4f}")
print(f"  Exposure₀ : {exposure_0:,.2f} active units")
print("-" * 70)
print(f"  {'Scenario':<28} {'Mean':>14} {'P90':>14} {'P95':>14} {'P99':>14}")
print(f"  {'':28} {'Nominal':>14} {'Nominal':>14} {'Nominal':>14} {'Nominal':>14}")
for scen, res in results.items():
    tot = res["nominal"].sum(axis=1)
    print(f"  {scen:<28} {tot.mean():>14,.0f} {np.percentile(tot,90):>14,.0f} "
          f"{np.percentile(tot,95):>14,.0f} {np.percentile(tot,99):>14,.0f}")
print()
print(f"  {'Scenario':<28} {'Mean Disc':>14} {'P90 Disc':>14} {'P95 Disc':>14} {'P99 Disc':>14}")
for scen, res in results.items():
    tot = res["discounted"].sum(axis=1)
    print(f"  {scen:<28} {tot.mean():>14,.0f} {np.percentile(tot,90):>14,.0f} "
          f"{np.percentile(tot,95):>14,.0f} {np.percentile(tot,99):>14,.0f}")
print("-" * 70)
print(f"  Output → {OUTPUT_PATH}")
print("-" * 70)
print("  This simulation represents loss uncertainty around the Step 3")
print("  deterministic baseline. Results are suitable for risk analysis,")
print("  capital assessment, and stress testing.")
print("=" * 70)